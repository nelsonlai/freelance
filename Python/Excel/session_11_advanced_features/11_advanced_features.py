"""
Session 11: Advanced Excel Features - Comprehensive Examples
============================================================

This script demonstrates advanced Excel features including working with images,
hyperlinks, comments, and integration with other file formats.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from openpyxl.styles import Font, colors
from pathlib import Path
from datetime import datetime
import json
import sqlite3

print("=" * 80)
print("Session 11: Advanced Excel Features")
print("=" * 80)
print()

# Create output directory
output_dir = Path(__file__).parent / 'sample_files'
output_dir.mkdir(exist_ok=True)

# ==============================================================================
# 1. WORKING WITH HYPERLINKS
# ==============================================================================

print("1. WORKING WITH HYPERLINKS")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Links"

# Web hyperlink
ws['A1'] = 'Visit Python.org'
ws['A1'].hyperlink = 'https://www.python.org'
ws['A1'].font = Font(color=colors.BLUE, underline='single')

# Internal sheet link
ws2 = wb.create_sheet('Data')
ws2['A1'] = 'This is the data sheet'
ws['A2'] = 'Go to Data Sheet'
ws['A2'].hyperlink = '#Data!A1'
ws['A2'].font = Font(color=colors.BLUE, underline='single')

# Email hyperlink
ws['A3'] = 'Send Email'
ws['A3'].hyperlink = 'mailto:info@example.com'
ws['A3'].font = Font(color=colors.BLUE, underline='single')

# File hyperlink
ws['A4'] = 'Open Document'
ws['A4'].hyperlink = 'file:///path/to/document.pdf'
ws['A4'].font = Font(color=colors.BLUE, underline='single')

hyperlinks_file = output_dir / '01_hyperlinks.xlsx'
wb.save(hyperlinks_file)
print(f"✓ Created: {hyperlinks_file.name}")
print(f"  4 different types of hyperlinks added")
print()

# ==============================================================================
# 2. WORKING WITH COMMENTS
# ==============================================================================

print("2. WORKING WITH COMMENTS")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add data with comments
ws['A1'] = 'Revenue'
ws['A1'].comment = Comment(
    'Annual revenue target\nVerified by Finance Dept.',
    'Financial Controller'
)

ws['B1'] = 1000000
ws['B1'].comment = Comment('Target for Q1 2024', 'Manager')

ws['A2'] = 'Expenses'
ws['A2'].comment = Comment('Operating expenses only', 'Accountant')

ws['B2'] = 750000
ws['B2'].comment = Comment(
    'Breakdown:\n- Salaries: $500K\n- Rent: $150K\n- Other: $100K',
    'CFO'
)

# Multi-line comment
ws['A4'] = 'Important Note'
ws['A4'].comment = Comment(
    'This is a multi-line comment.\n'
    'Line 2: Additional information\n'
    'Line 3: Contact John for questions\n'
    'Last updated: ' + datetime.now().strftime('%Y-%m-%d'),
    'System Admin'
)

comments_file = output_dir / '02_comments.xlsx'
wb.save(comments_file)
print(f"✓ Created: {comments_file.name}")
print(f"  5 cells have comments with detailed information")
print()

# ==============================================================================
# 3. DOCUMENT PROPERTIES AND METADATA
# ==============================================================================

print("3. DOCUMENT PROPERTIES AND METADATA")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Set comprehensive document properties
wb.properties.title = 'Q1 2024 Financial Report'
wb.properties.subject = 'Financial Analysis'
wb.properties.creator = 'Python Automation System'
wb.properties.keywords = 'finance, Q1, 2024, report, analysis'
wb.properties.description = 'Comprehensive financial report for Q1 2024 including revenue, expenses, and projections'
wb.properties.category = 'Financial Reports'
wb.properties.company = 'Acme Corporation'
wb.properties.manager = 'Jane Smith'
wb.properties.created = datetime(2024, 1, 1)
wb.properties.modified = datetime.now()
wb.properties.lastModifiedBy = 'Automation Script v2.0'

# Add content
ws['A1'] = 'Financial Report'
ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

metadata_file = output_dir / '03_metadata.xlsx'
wb.save(metadata_file)
print(f"✓ Created: {metadata_file.name}")
print(f"  Complete document properties set:")
print(f"    Title: {wb.properties.title}")
print(f"    Company: {wb.properties.company}")
print(f"    Category: {wb.properties.category}")
print()

# ==============================================================================
# 4. EXCEL TO CSV CONVERSION
# ==============================================================================

print("4. EXCEL TO CSV CONVERSION")
print("-" * 80)

# Create sample Excel file
df_sample = pd.DataFrame({
    'Product': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Webcam'],
    'Price': [1000, 25, 75, 300, 90],
    'Stock': [50, 200, 150, 75, 100]
})

excel_file = output_dir / 'products.xlsx'
df_sample.to_excel(excel_file, index=False)
print(f"✓ Created Excel file: {excel_file.name}")

# Convert to CSV
csv_file = output_dir / 'products.csv'
df_sample.to_csv(csv_file, index=False)
print(f"✓ Converted to CSV: {csv_file.name}")

# Compare file sizes
excel_size = excel_file.stat().st_size / 1024
csv_size = csv_file.stat().st_size / 1024
print(f"  Excel size: {excel_size:.2f} KB")
print(f"  CSV size: {csv_size:.2f} KB")
print(f"  CSV is {excel_size/csv_size:.1f}x smaller")
print()

# ==============================================================================
# 5. EXCEL TO JSON CONVERSION
# ==============================================================================

print("5. EXCEL TO JSON CONVERSION")
print("-" * 80)

# Read Excel data
df = pd.read_excel(excel_file)

# Convert to different JSON formats
json_records_file = output_dir / 'products_records.json'
json_records = df.to_json(orient='records', indent=2)
with open(json_records_file, 'w') as f:
    f.write(json_records)
print(f"✓ Created JSON (records format): {json_records_file.name}")

json_columns_file = output_dir / 'products_columns.json'
json_columns = df.to_json(orient='columns', indent=2)
with open(json_columns_file, 'w') as f:
    f.write(json_columns)
print(f"✓ Created JSON (columns format): {json_columns_file.name}")

print(f"\nSample JSON (records format):")
print(json.dumps(json.loads(json_records)[:2], indent=2))
print()

# ==============================================================================
# 6. JSON TO EXCEL CONVERSION
# ==============================================================================

print("6. JSON TO EXCEL CONVERSION")
print("-" * 80)

# Create sample JSON data
json_data = [
    {'name': 'Alice', 'age': 30, 'city': 'NYC', 'salary': 75000},
    {'name': 'Bob', 'age': 25, 'city': 'LA', 'salary': 65000},
    {'name': 'Charlie', 'age': 35, 'city': 'Chicago', 'salary': 80000}
]

json_input_file = output_dir / 'employees.json'
with open(json_input_file, 'w') as f:
    json.dump(json_data, f, indent=2)
print(f"✓ Created JSON file: {json_input_file.name}")

# Convert JSON to Excel
with open(json_input_file, 'r') as f:
    data = json.load(f)

df_from_json = pd.DataFrame(data)
excel_from_json = output_dir / 'employees_from_json.xlsx'
df_from_json.to_excel(excel_from_json, index=False)
print(f"✓ Converted to Excel: {excel_from_json.name}")
print()

# ==============================================================================
# 7. EXCEL TO SQL DATABASE
# ==============================================================================

print("7. EXCEL TO SQL DATABASE")
print("-" * 80)

# Create SQLite database
db_file = output_dir / 'company.db'
conn = sqlite3.connect(db_file)

# Read Excel and write to SQL
df_products = pd.read_excel(excel_file)
df_products.to_sql('products', conn, if_exists='replace', index=False)
print(f"✓ Exported to SQLite: {db_file.name}")
print(f"  Table: products ({len(df_products)} rows)")

# Add another table
df_employees = pd.read_excel(excel_from_json)
df_employees.to_sql('employees', conn, if_exists='replace', index=False)
print(f"  Table: employees ({len(df_employees)} rows)")

# Verify by reading back
df_verify = pd.read_sql("SELECT * FROM products", conn)
print(f"\n Verification (first 3 rows):")
print(df_verify.head(3).to_string(index=False))
print()

conn.close()

# ==============================================================================
# 8. SQL TO EXCEL EXPORT
# ==============================================================================

print("8. SQL TO EXCEL EXPORT")
print("-" * 80)

# Reconnect to database
conn = sqlite3.connect(db_file)

# Export multiple tables to one Excel file
sql_to_excel_file = output_dir / 'database_export.xlsx'

with pd.ExcelWriter(sql_to_excel_file) as writer:
    # Products table
    df_products = pd.read_sql("SELECT * FROM products", conn)
    df_products.to_excel(writer, sheet_name='Products', index=False)
    
    # Employees table
    df_employees = pd.read_sql("SELECT * FROM employees", conn)
    df_employees.to_excel(writer, sheet_name='Employees', index=False)
    
    # Custom query
    query = """
    SELECT Product, Price, Stock,
           Price * Stock as Total_Value
    FROM products
    ORDER BY Total_Value DESC
    """
    df_analysis = pd.read_sql(query, conn)
    df_analysis.to_excel(writer, sheet_name='Value_Analysis', index=False)

conn.close()

print(f"✓ Exported database to Excel: {sql_to_excel_file.name}")
print(f"  3 sheets: Products, Employees, Value_Analysis")
print()

# ==============================================================================
# 9. UNIVERSAL FILE CONVERTER
# ==============================================================================

print("9. UNIVERSAL FILE CONVERTER")
print("-" * 80)

def convert_file(input_file, output_format):
    """Convert between different formats"""
    input_path = Path(input_file)
    input_format = input_path.suffix.lower().strip('.')
    
    print(f"Converting {input_path.name} from {input_format} to {output_format}")
    
    # Read file
    if input_format == 'xlsx':
        df = pd.read_excel(input_file)
    elif input_format == 'csv':
        df = pd.read_csv(input_file)
    elif input_format == 'json':
        df = pd.read_json(input_file)
    else:
        print(f"  ✗ Unsupported input format: {input_format}")
        return None
    
    # Write file
    output_file = input_path.stem + f'.{output_format}'
    output_path = input_path.parent / output_file
    
    if output_format == 'xlsx':
        df.to_excel(output_path, index=False)
    elif output_format == 'csv':
        df.to_csv(output_path, index=False)
    elif output_format == 'json':
        df.to_json(output_path, orient='records', indent=2)
    else:
        print(f"  ✗ Unsupported output format: {output_format}")
        return None
    
    print(f"  ✓ Converted to: {output_file}")
    return output_path

# Test conversions
print("Testing conversions:")
convert_file(excel_file, 'json')
convert_file(csv_file, 'xlsx')
convert_file(json_input_file, 'csv')
print()

# ==============================================================================
# 10. COMPREHENSIVE MULTI-FORMAT WORKFLOW
# ==============================================================================

print("10. COMPREHENSIVE MULTI-FORMAT WORKFLOW")
print("-" * 80)

def multi_format_workflow():
    """Complete workflow using multiple formats"""
    
    # Step 1: Create data in Python
    print("Step 1: Creating data...")
    sales_data = {
        'Date': pd.date_range('2024-01-01', periods=10),
        'Product': ['Laptop'] * 5 + ['Mouse'] * 5,
        'Sales': np.random.randint(1000, 5000, 10),
        'Region': ['North', 'South'] * 5
    }
    df = pd.DataFrame(sales_data)
    print(f"  ✓ Created DataFrame: {len(df)} rows")
    
    # Step 2: Save as Excel
    print("\nStep 2: Saving as Excel...")
    excel_output = output_dir / 'workflow_data.xlsx'
    df.to_excel(excel_output, index=False)
    print(f"  ✓ Saved: {excel_output.name}")
    
    # Step 3: Convert to CSV (for sharing)
    print("\nStep 3: Converting to CSV...")
    csv_output = output_dir / 'workflow_data.csv'
    df.to_csv(csv_output, index=False)
    print(f"  ✓ Saved: {csv_output.name}")
    
    # Step 4: Convert to JSON (for API)
    print("\nStep 4: Converting to JSON...")
    json_output = output_dir / 'workflow_data.json'
    df.to_json(json_output, orient='records', indent=2)
    print(f"  ✓ Saved: {json_output.name}")
    
    # Step 5: Store in SQLite (for querying)
    print("\nStep 5: Storing in database...")
    workflow_db = output_dir / 'workflow.db'
    conn = sqlite3.connect(workflow_db)
    df.to_sql('sales', conn, if_exists='replace', index=False)
    print(f"  ✓ Saved to: {workflow_db.name}")
    
    # Step 6: Query and export
    print("\nStep 6: Querying and exporting...")
    query = """
    SELECT Product, SUM(Sales) as Total_Sales, COUNT(*) as Count
    FROM sales
    GROUP BY Product
    """
    df_summary = pd.read_sql(query, conn)
    conn.close()
    
    summary_excel = output_dir / 'workflow_summary.xlsx'
    df_summary.to_excel(summary_excel, index=False)
    print(f"  ✓ Summary saved: {summary_excel.name}")
    print(f"\n  Summary:")
    print(df_summary.to_string(index=False, justify='left'))
    
    print("\n  ✓ Workflow complete!")
    return True

# Run comprehensive workflow
multi_format_workflow()
print()

# ==============================================================================
# 11. BATCH FORMAT CONVERSION
# ==============================================================================

print("11. BATCH FORMAT CONVERSION")
print("-" * 80)

def batch_convert(directory, from_format, to_format):
    """Convert all files of one format to another"""
    files = list(Path(directory).glob(f'*.{from_format}'))
    
    if not files:
        print(f"No .{from_format} files found")
        return
    
    print(f"Converting {len(files)} files from {from_format} to {to_format}")
    
    for file in files:
        try:
            # Read
            if from_format == 'xlsx':
                df = pd.read_excel(file)
            elif from_format == 'csv':
                df = pd.read_csv(file)
            elif from_format == 'json':
                df = pd.read_json(file)
            
            # Write
            output_file = file.stem + f'.{to_format}'
            output_path = file.parent / output_file
            
            if to_format == 'xlsx':
                df.to_excel(output_path, index=False)
            elif to_format == 'csv':
                df.to_csv(output_path, index=False)
            elif to_format == 'json':
                df.to_json(output_path, orient='records', indent=2)
            
            print(f"  ✓ {file.name} → {output_file}")
            
        except Exception as e:
            print(f"  ✗ {file.name}: {e}")

# Test batch conversion
print("Converting all JSON files to Excel:")
batch_convert(output_dir, 'json', 'xlsx')
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 80)
print("SESSION 11 COMPLETE!")
print("=" * 80)
print()
print("What you learned:")
print("✓ Add hyperlinks (web, internal, email, file)")
print("✓ Work with cell comments")
print("✓ Set document properties and metadata")
print("✓ Convert Excel to CSV")
print("✓ Convert Excel to/from JSON")
print("✓ Export Excel to SQL database")
print("✓ Import SQL data to Excel")
print("✓ Build universal file converter")
print("✓ Create multi-format workflows")
print("✓ Batch convert between formats")
print()
print("Files created:")
for file in sorted(output_dir.glob('*')):
    if file.is_file():
        print(f"  - {file.name}")
print()
print("Format Integration Summary:")
print("  ✓ Excel ↔ CSV: Fast, simple, widely compatible")
print("  ✓ Excel ↔ JSON: Structured data, API integration")
print("  ✓ Excel ↔ SQL: Queryable, relational data")
print("  ✓ Universal converter: Handle any format")
print()
print("Next: Session 12 - Real-World Projects")
print("=" * 80)

