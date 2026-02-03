"""
Session 1: Setup and Basics - Python Excel Libraries
=====================================================

This script demonstrates the fundamental concepts of working with Excel files
using Python. Run each section to see the examples in action.

Prerequisites:
- pandas, openpyxl, xlsxwriter, numpy installed
- Sample Excel files in the sample_files directory
"""

import pandas as pd
import numpy as np
from pathlib import Path
import sys

# Add parent directory to path to access sample_data
sys.path.append(str(Path(__file__).parent.parent))

print("=" * 70)
print("Session 1: Python Excel Libraries - Setup and Basics")
print("=" * 70)
print()

# ==============================================================================
# 1. VERIFY LIBRARY INSTALLATIONS
# ==============================================================================

print("1. VERIFYING LIBRARY INSTALLATIONS")
print("-" * 70)

try:
    import pandas as pd
    import openpyxl
    import xlsxwriter
    import numpy as np
    
    print(f"✓ pandas version: {pd.__version__}")
    print(f"✓ openpyxl version: {openpyxl.__version__}")
    print(f"✓ xlsxwriter version: {xlsxwriter.__version__}")
    print(f"✓ numpy version: {np.__version__}")
    print("\n✓ All required libraries are installed!")
    
except ImportError as e:
    print(f"✗ Error: {e}")
    print("Please install missing packages:")
    print("pip install pandas openpyxl xlsxwriter numpy")
    sys.exit(1)

print()

# ==============================================================================
# 2. CREATE SAMPLE EXCEL FILE
# ==============================================================================

print("2. CREATING SAMPLE EXCEL FILE")
print("-" * 70)

# Create sample data
sample_data = {
    'Product': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Webcam'],
    'Category': ['Electronics', 'Accessories', 'Accessories', 'Electronics', 'Electronics'],
    'Price': [999.99, 29.99, 79.99, 299.99, 89.99],
    'Quantity': [50, 200, 150, 75, 100],
    'Supplier': ['TechCorp', 'GadgetCo', 'GadgetCo', 'TechCorp', 'TechCorp']
}

df_sample = pd.DataFrame(sample_data)
print("Sample DataFrame created:")
print(df_sample)
print()

# Save to Excel
sample_file = Path(__file__).parent / 'sample_files' / 'sample_products.xlsx'
sample_file.parent.mkdir(exist_ok=True)
df_sample.to_excel(sample_file, index=False)
print(f"✓ Sample file created: {sample_file}")
print()

# ==============================================================================
# 3. READING EXCEL FILES WITH PANDAS
# ==============================================================================

print("3. READING EXCEL FILES WITH PANDAS")
print("-" * 70)

# Read the Excel file we just created
df_read = pd.read_excel(sample_file)
print("File read successfully using pandas!")
print()

print("First few rows:")
print(df_read.head())
print()

print("DataFrame Info:")
print(df_read.info())
print()

print("Statistical Summary:")
print(df_read.describe())
print()

# ==============================================================================
# 4. BASIC DATAFRAME OPERATIONS
# ==============================================================================

print("4. BASIC DATAFRAME OPERATIONS")
print("-" * 70)

# Select a column
print("Selecting 'Product' column:")
print(df_read['Product'])
print()

# Select multiple columns
print("Selecting 'Product' and 'Price' columns:")
print(df_read[['Product', 'Price']])
print()

# Filter rows
print("Products with Price > 100:")
expensive = df_read[df_read['Price'] > 100]
print(expensive)
print()

# Add a new column
df_read['Total_Value'] = df_read['Price'] * df_read['Quantity']
print("DataFrame with new 'Total_Value' column:")
print(df_read)
print()

# Sort data
print("Sorted by Price (descending):")
df_sorted = df_read.sort_values('Price', ascending=False)
print(df_sorted[['Product', 'Price']])
print()

# Group and aggregate
print("Total Quantity by Category:")
category_summary = df_read.groupby('Category')['Quantity'].sum()
print(category_summary)
print()

# ==============================================================================
# 5. READING WITH OPENPYXL
# ==============================================================================

print("5. READING EXCEL FILES WITH OPENPYXL")
print("-" * 70)

from openpyxl import load_workbook

# Load workbook
wb = load_workbook(sample_file)
print(f"✓ Workbook loaded: {sample_file}")
print()

# Get sheet names
print(f"Sheet names: {wb.sheetnames}")
print()

# Access active sheet
ws = wb.active
print(f"Active sheet: {ws.title}")
print()

# Read cell values
print("Reading cell values:")
print(f"Cell A1: {ws['A1'].value}")
print(f"Cell B2: {ws['B2'].value}")
print(f"Cell C3 (using row/col): {ws.cell(row=3, column=3).value}")
print()

# Get dimensions
print(f"Sheet dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")
print()

# Iterate through rows
print("First 5 rows:")
for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    print(f"Row {row_num}: {row}")
print()

# ==============================================================================
# 6. CREATING EXCEL FILES WITH PANDAS
# ==============================================================================

print("6. CREATING EXCEL FILES WITH PANDAS")
print("-" * 70)

# Create new data
new_data = {
    'Employee': ['Alice', 'Bob', 'Charlie', 'Diana', 'Eve'],
    'Department': ['Sales', 'IT', 'Sales', 'HR', 'IT'],
    'Salary': [60000, 75000, 65000, 58000, 72000],
    'Years': [3, 5, 2, 7, 4]
}

df_employees = pd.DataFrame(new_data)
print("Employee DataFrame:")
print(df_employees)
print()

# Write to Excel
output_file = Path(__file__).parent / 'sample_files' / 'employees.xlsx'
df_employees.to_excel(output_file, index=False, sheet_name='Employees')
print(f"✓ Employee data written to: {output_file}")
print()

# ==============================================================================
# 7. CREATING EXCEL FILES WITH MULTIPLE SHEETS
# ==============================================================================

print("7. CREATING EXCEL FILES WITH MULTIPLE SHEETS")
print("-" * 70)

# Create multiple DataFrames
df_sales = pd.DataFrame({
    'Month': ['Jan', 'Feb', 'Mar', 'Apr'],
    'Revenue': [50000, 55000, 52000, 58000],
    'Expenses': [30000, 32000, 31000, 33000]
})

df_customers = pd.DataFrame({
    'Customer': ['Acme Corp', 'TechStart', 'Global Inc', 'Local Biz'],
    'Total_Orders': [45, 32, 67, 23],
    'Total_Revenue': [125000, 89000, 178000, 56000]
})

# Write to Excel with multiple sheets
multi_sheet_file = Path(__file__).parent / 'sample_files' / 'business_report.xlsx'

with pd.ExcelWriter(multi_sheet_file, engine='openpyxl') as writer:
    df_sales.to_excel(writer, sheet_name='Sales', index=False)
    df_customers.to_excel(writer, sheet_name='Customers', index=False)
    df_employees.to_excel(writer, sheet_name='Employees', index=False)

print(f"✓ Multi-sheet workbook created: {multi_sheet_file}")
print(f"  Sheets: Sales, Customers, Employees")
print()

# ==============================================================================
# 8. READING SPECIFIC SHEETS
# ==============================================================================

print("8. READING SPECIFIC SHEETS")
print("-" * 70)

# Read specific sheet by name
df_sales_read = pd.read_excel(multi_sheet_file, sheet_name='Sales')
print("Sales sheet:")
print(df_sales_read)
print()

# Read specific sheet by index (0-based)
df_customers_read = pd.read_excel(multi_sheet_file, sheet_name=1)
print("Customers sheet (by index):")
print(df_customers_read)
print()

# Read all sheets at once
all_sheets = pd.read_excel(multi_sheet_file, sheet_name=None)
print(f"All sheets loaded: {list(all_sheets.keys())}")
print()

for sheet_name, df in all_sheets.items():
    print(f"\n{sheet_name} sheet preview:")
    print(df.head(2))
print()

# ==============================================================================
# 9. COMBINING PANDAS AND OPENPYXL
# ==============================================================================

print("9. COMBINING PANDAS AND OPENPYXL FOR FORMATTING")
print("-" * 70)

from openpyxl.styles import Font, PatternFill, Alignment

# Create data with pandas
summary_data = {
    'Region': ['North', 'South', 'East', 'West'],
    'Q1_Sales': [125000, 98000, 142000, 115000],
    'Q2_Sales': [132000, 105000, 138000, 121000],
    'Q3_Sales': [128000, 112000, 145000, 118000],
    'Q4_Sales': [145000, 118000, 155000, 130000]
}

df_summary = pd.DataFrame(summary_data)
df_summary['Total'] = df_summary[['Q1_Sales', 'Q2_Sales', 'Q3_Sales', 'Q4_Sales']].sum(axis=1)
print("Summary DataFrame:")
print(df_summary)
print()

# Write with pandas
formatted_file = Path(__file__).parent / 'sample_files' / 'formatted_summary.xlsx'
df_summary.to_excel(formatted_file, index=False, sheet_name='Summary')

# Format with openpyxl
wb = load_workbook(formatted_file)
ws = wb['Summary']

# Format header row
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max_length + 3

wb.save(formatted_file)
print(f"✓ Formatted summary created: {formatted_file}")
print()

# ==============================================================================
# 10. ERROR HANDLING
# ==============================================================================

print("10. ERROR HANDLING BEST PRACTICES")
print("-" * 70)

# Example 1: File not found
print("Example 1: Handling FileNotFoundError")
try:
    df = pd.read_excel('non_existent_file.xlsx')
except FileNotFoundError:
    print("✓ Caught: File not found error handled gracefully")
print()

# Example 2: Invalid sheet name
print("Example 2: Handling invalid sheet name")
try:
    df = pd.read_excel(multi_sheet_file, sheet_name='InvalidSheet')
except ValueError as e:
    print(f"✓ Caught: {e}")
print()

# Example 3: Checking file existence
print("Example 3: Checking file existence before reading")
file_to_check = Path('maybe_exists.xlsx')
if file_to_check.exists():
    df = pd.read_excel(file_to_check)
    print("File exists and loaded successfully")
else:
    print(f"✓ File does not exist: {file_to_check}")
print()

# ==============================================================================
# 11. PERFORMANCE COMPARISON
# ==============================================================================

print("11. PERFORMANCE COMPARISON: PANDAS VS OPENPYXL")
print("-" * 70)

import time

# Create a larger dataset
large_data = {
    'ID': range(1, 1001),
    'Name': [f'Product_{i}' for i in range(1, 1001)],
    'Value': np.random.randint(10, 1000, 1000)
}

df_large = pd.DataFrame(large_data)
large_file = Path(__file__).parent / 'sample_files' / 'large_dataset.xlsx'
df_large.to_excel(large_file, index=False)
print(f"✓ Large dataset created: 1000 rows")
print()

# Test pandas reading speed
start_time = time.time()
df_pandas = pd.read_excel(large_file)
pandas_time = time.time() - start_time
print(f"Pandas read time: {pandas_time:.4f} seconds")

# Test openpyxl reading speed
start_time = time.time()
wb_openpyxl = load_workbook(large_file)
ws_openpyxl = wb_openpyxl.active
data_openpyxl = [[cell.value for cell in row] for row in ws_openpyxl.iter_rows()]
openpyxl_time = time.time() - start_time
print(f"Openpyxl read time: {openpyxl_time:.4f} seconds")
print()

if pandas_time < openpyxl_time:
    speedup = openpyxl_time / pandas_time
    print(f"✓ Pandas is {speedup:.2f}x faster for reading data")
else:
    speedup = pandas_time / openpyxl_time
    print(f"✓ Openpyxl is {speedup:.2f}x faster for reading data")
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 70)
print("SESSION 1 COMPLETE!")
print("=" * 70)
print()
print("What you learned:")
print("✓ Verified Python Excel library installations")
print("✓ Created and read Excel files with pandas")
print("✓ Performed basic DataFrame operations")
print("✓ Read Excel files with openpyxl")
print("✓ Created multi-sheet Excel workbooks")
print("✓ Combined pandas and openpyxl for formatted output")
print("✓ Implemented error handling best practices")
print("✓ Compared performance of different libraries")
print()
print("Files created:")
print(f"  - {sample_file}")
print(f"  - {output_file}")
print(f"  - {multi_sheet_file}")
print(f"  - {formatted_file}")
print(f"  - {large_file}")
print()
print("Next: Proceed to Session 2 for advanced reading techniques!")
print("=" * 70)

