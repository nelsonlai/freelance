"""
Session 6: Working with Formulas and Charts - Comprehensive Examples
=====================================================================

This script demonstrates comprehensive techniques for adding formulas
and creating charts in Excel files.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    LineChart, BarChart, PieChart, ScatterChart,
    Reference, Series
)
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart.marker import Marker
import xlsxwriter
from pathlib import Path

print("=" * 80)
print("Session 6: Working with Formulas and Charts")
print("=" * 80)
print()

# Create output directory
output_dir = Path(__file__).parent / 'sample_files'
output_dir.mkdir(exist_ok=True)

# ==============================================================================
# 1. BASIC FORMULAS
# ==============================================================================

print("1. BASIC FORMULAS")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Basic Formulas"

# Add data
ws['A1'] = 'Price'
ws['B1'] = 'Quantity'
ws['C1'] = 'Total'

prices = [100, 200, 300, 150, 250]
quantities = [5, 3, 7, 10, 4]

for row, (price, qty) in enumerate(zip(prices, quantities), 2):
    ws[f'A{row}'] = price
    ws[f'B{row}'] = qty
    ws[f'C{row}'] = f'=A{row}*B{row}'  # Formula for total

# Add sum formula
last_row = len(prices) + 2
ws[f'A{last_row}'] = 'TOTAL:'
ws[f'C{last_row}'] = f'=SUM(C2:C{last_row-1})'
ws[f'C{last_row}'].font = Font(bold=True)

formulas_file = output_dir / '01_basic_formulas.xlsx'
wb.save(formulas_file)
print(f"✓ Created: {formulas_file.name}")
print()

# ==============================================================================
# 2. ADVANCED FORMULAS
# ==============================================================================

print("2. ADVANCED FORMULAS (IF, SUMIF, VLOOKUP)")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Advanced Formulas"

# Add data with headers
ws['A1'] = 'Product'
ws['B1'] = 'Sales'
ws['C1'] = 'Status'
ws['D1'] = 'Bonus'

products = [
    ('Laptop', 12000),
    ('Mouse', 3000),
    ('Keyboard', 4500),
    ('Monitor', 8500),
    ('Webcam', 6000)
]

for row, (product, sales) in enumerate(products, 2):
    ws[f'A{row}'] = product
    ws[f'B{row}'] = sales
    # IF formula: High if > 5000, Low otherwise
    ws[f'C{row}'] = f'=IF(B{row}>5000,"High","Low")'
    # Bonus calculation: 10% if High, 5% if Low
    ws[f'D{row}'] = f'=IF(C{row}="High",B{row}*0.1,B{row}*0.05)'

# Add summary statistics
ws['F1'] = 'Summary'
ws['F2'] = 'High Sales Count:'
ws['G2'] = f'=COUNTIF(C2:C6,"High")'

ws['F3'] = 'Low Sales Count:'
ws['G3'] = f'=COUNTIF(C2:C6,"Low")'

ws['F4'] = 'High Sales Total:'
ws['G4'] = f'=SUMIF(C2:C6,"High",B2:B6)'

ws['F5'] = 'Average Sales:'
ws['G5'] = '=AVERAGE(B2:B6)'

advanced_formulas_file = output_dir / '02_advanced_formulas.xlsx'
wb.save(advanced_formulas_file)
print(f"✓ Created: {advanced_formulas_file.name}")
print()

# ==============================================================================
# 3. COLUMN CHART
# ==============================================================================

print("3. COLUMN CHART")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Sales Chart"

# Add data
data = [
    ['Product', 'Sales'],
    ['Laptop', 120000],
    ['Mouse', 35000],
    ['Keyboard', 45000],
    ['Monitor', 78000],
    ['Webcam', 52000]
]

for row in data:
    ws.append(row)

# Create column chart
chart = BarChart()
chart.type = 'col'
chart.title = 'Product Sales'
chart.x_axis.title = 'Product'
chart.y_axis.title = 'Sales ($)'
chart.height = 12
chart.width = 20
chart.style = 10

# Define data ranges
data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
categories_ref = Reference(ws, min_col=1, min_row=2, max_row=6)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(categories_ref)

# Add chart to worksheet
ws.add_chart(chart, 'D2')

column_chart_file = output_dir / '03_column_chart.xlsx'
wb.save(column_chart_file)
print(f"✓ Created: {column_chart_file.name}")
print()

# ==============================================================================
# 4. LINE CHART (TREND)
# ==============================================================================

print("4. LINE CHART (TREND)")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Monthly Trend"

# Add monthly data
data = [
    ['Month', 'Sales', 'Costs', 'Profit'],
    ['Jan', 45000, 30000, 15000],
    ['Feb', 52000, 34000, 18000],
    ['Mar', 48000, 32000, 16000],
    ['Apr', 61000, 39000, 22000],
    ['May', 55000, 36000, 19000],
    ['Jun', 58000, 38000, 20000]
]

for row in data:
    ws.append(row)

# Create line chart with multiple series
chart = LineChart()
chart.title = 'Monthly Sales Analysis'
chart.style = 12
chart.x_axis.title = 'Month'
chart.y_axis.title = 'Amount ($)'
chart.height = 12
chart.width = 20

# Y-axis values
y_values = Reference(ws, min_col=2, min_row=1, max_row=7)
# X-axis categories
x_categories = Reference(ws, min_col=1, min_row=2, max_row=7)

# Add data (Sales, Costs, Profit)
data_ref = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=7)
categories_ref = Reference(ws, min_col=1, min_row=2, max_row=7)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(categories_ref)

ws.add_chart(chart, 'F2')

line_chart_file = output_dir / '04_line_chart.xlsx'
wb.save(line_chart_file)
print(f"✓ Created: {line_chart_file.name}")
print()

# ==============================================================================
# 5. PIE CHART
# ==============================================================================

print("5. PIE CHART")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Budget Distribution"

# Add data
data = [
    ['Department', 'Budget'],
    ['Sales', 500000],
    ['Marketing', 200000],
    ['IT', 350000],
    ['HR', 150000],
    ['Operations', 400000]
]

for row in data:
    ws.append(row)

# Create pie chart
chart = PieChart()
chart.title = 'Budget Distribution by Department'
chart.height = 12
chart.width = 20

data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
labels_ref = Reference(ws, min_col=1, min_row=2, max_row=6)

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(labels_ref)

ws.add_chart(chart, 'D2')

pie_chart_file = output_dir / '05_pie_chart.xlsx'
wb.save(pie_chart_file)
print(f"✓ Created: {pie_chart_file.name}")
print()

# ==============================================================================
# 6. SCATTER PLOT
# ==============================================================================

print("6. SCATTER PLOT (CORRELATION)")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Price Analysis"

# Add data
ws.append(['Price', 'Units Sold'])
data = [(10, 100), (15, 85), (20, 70), (25, 60), (30, 45), (35, 35), (40, 28)]
for price, units in data:
    ws.append([price, units])

# Create scatter chart
chart = ScatterChart()
chart.title = 'Price vs Units Sold'
chart.style = 13
chart.x_axis.title = 'Price ($)'
chart.y_axis.title = 'Units Sold'
chart.height = 12
chart.width = 18

xvalues = Reference(ws, min_col=1, min_row=2, max_row=8)
yvalues = Reference(ws, min_col=2, min_row=2, max_row=8)

series = Series(yvalues, xvalues, title='Sales Data')
chart.series.append(series)

ws.add_chart(chart, 'D2')

scatter_chart_file = output_dir / '06_scatter_chart.xlsx'
wb.save(scatter_chart_file)
print(f"✓ Created: {scatter_chart_file.name}")
print()

# ==============================================================================
# 7. XLSXWRITER CHARTS
# ==============================================================================

print("7. XLSXWRITER CHARTS")
print("-" * 80)

xlsxwriter_file = output_dir / '07_xlsxwriter_charts.xlsx'
workbook = xlsxwriter.Workbook(xlsxwriter_file)
worksheet = workbook.add_worksheet('Quarterly Sales')

# Add data
headers = ['Product', 'Q1', 'Q2', 'Q3', 'Q4']
worksheet.write_row('A1', headers, workbook.add_format({'bold': True}))

data = [
    ['Laptop', 45, 52, 48, 56],
    ['Mouse', 120, 135, 128, 142],
    ['Keyboard', 85, 90, 88, 95],
    ['Monitor', 32, 38, 35, 41]
]

for row_idx, row_data in enumerate(data, 1):
    worksheet.write_row(row_idx, 0, row_data)

# Create column chart
chart = workbook.add_chart({'type': 'column'})

# Add series for each product
for row in range(1, 5):
    chart.add_series({
        'name': ['Quarterly Sales', row, 0],
        'categories': ['Quarterly Sales', 0, 1, 0, 4],
        'values': ['Quarterly Sales', row, 1, row, 4]
    })

# Configure chart
chart.set_title({'name': 'Quarterly Product Sales'})
chart.set_x_axis({'name': 'Quarter'})
chart.set_y_axis({'name': 'Units Sold'})
chart.set_style(11)
chart.set_size({'width': 720, 'height': 576})

# Insert chart
worksheet.insert_chart('G2', chart)

workbook.close()
print(f"✓ Created: {xlsxwriter_file.name}")
print()

# ==============================================================================
# 8. COMPREHENSIVE DASHBOARD
# ==============================================================================

print("8. COMPREHENSIVE DASHBOARD")
print("-" * 80)

# Create data with pandas
df = pd.DataFrame({
    'Month': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
    'Revenue': [45000, 52000, 48000, 61000, 55000, 58000],
    'Costs': [30000, 34000, 32000, 39000, 36000, 38000],
    'Marketing': [5000, 6000, 5500, 7000, 6500, 6800]
})

df['Profit'] = df['Revenue'] - df['Costs']
df['Profit_Margin_%'] = (df['Profit'] / df['Revenue'] * 100).round(1)

# Write to Excel
dashboard_file = output_dir / '08_comprehensive_dashboard.xlsx'
with pd.ExcelWriter(dashboard_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Data', index=False)

# Load and enhance
wb = load_workbook(dashboard_file)
ws = wb['Data']

# Add totals row with formulas
last_row = len(df) + 2
ws[f'A{last_row}'] = 'TOTAL'
ws[f'B{last_row}'] = f'=SUM(B2:B{last_row-1})'
ws[f'C{last_row}'] = f'=SUM(C2:C{last_row-1})'
ws[f'D{last_row}'] = f'=SUM(D2:D{last_row-1})'
ws[f'E{last_row}'] = f'=SUM(E2:E{last_row-1})'

# Format totals row
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws[f'{col}{last_row}']
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# Add KPIs section
ws['H1'] = 'Key Performance Indicators'
ws['H1'].font = Font(bold=True, size=14)

ws['H3'] = 'Average Revenue:'
ws['I3'] = f'=AVERAGE(B2:B{last_row-1})'
ws['I3'].number_format = '$#,##0'

ws['H4'] = 'Average Profit:'
ws['I4'] = f'=AVERAGE(E2:E{last_row-1})'
ws['I4'].number_format = '$#,##0'

ws['H5'] = 'Avg Profit Margin:'
ws['I5'] = f'=AVERAGE(F2:F{last_row-1})'
ws['I5'].number_format = '0.0"%"'

ws['H6'] = 'Best Month:'
ws['I6'] = f'=INDEX(A2:A{last_row-1},MATCH(MAX(B2:B{last_row-1}),B2:B{last_row-1},0))'

# Create line chart for revenue/costs
chart1 = LineChart()
chart1.title = 'Revenue vs Costs Trend'
chart1.style = 12
chart1.x_axis.title = 'Month'
chart1.y_axis.title = 'Amount ($)'
chart1.height = 10
chart1.width = 15

data_ref = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=7)
categories_ref = Reference(ws, min_col=1, min_row=2, max_row=7)

chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(categories_ref)

ws.add_chart(chart1, 'H8')

# Create bar chart for profit
chart2 = BarChart()
chart2.type = 'col'
chart2.title = 'Monthly Profit'
chart2.style = 11
chart2.x_axis.title = 'Month'
chart2.y_axis.title = 'Profit ($)'
chart2.height = 10
chart2.width = 15

profit_ref = Reference(ws, min_col=5, min_row=1, max_row=7)
chart2.add_data(profit_ref, titles_from_data=True)
chart2.set_categories(categories_ref)

ws.add_chart(chart2, 'H23')

wb.save(dashboard_file)
print(f"✓ Created: {dashboard_file.name}")
print()

# ==============================================================================
# 9. MULTI-CHART REPORT
# ==============================================================================

print("9. MULTI-CHART REPORT")
print("-" * 80)

# Create comprehensive sales data
np.random.seed(42)

regions = ['North', 'South', 'East', 'West']
products = ['Laptop', 'Mouse', 'Keyboard', 'Monitor']

sales_data = []
for region in regions:
    for product in products:
        sales_data.append({
            'Region': region,
            'Product': product,
            'Q1': np.random.randint(20, 80),
            'Q2': np.random.randint(20, 80),
            'Q3': np.random.randint(20, 80),
            'Q4': np.random.randint(20, 80)
        })

df_sales = pd.DataFrame(sales_data)
df_sales['Total'] = df_sales[['Q1', 'Q2', 'Q3', 'Q4']].sum(axis=1)

# Write to Excel
multi_chart_file = output_dir / '09_multi_chart_report.xlsx'
with pd.ExcelWriter(multi_chart_file, engine='openpyxl') as writer:
    df_sales.to_excel(writer, sheet_name='Sales Data', index=False)
    
    # Regional summary
    regional_summary = df_sales.groupby('Region')['Total'].sum().reset_index()
    regional_summary.to_excel(writer, sheet_name='Regional Summary', index=False)
    
    # Product summary
    product_summary = df_sales.groupby('Product')['Total'].sum().reset_index()
    product_summary.to_excel(writer, sheet_name='Product Summary', index=False)

# Load and add charts
wb = load_workbook(multi_chart_file)

# Chart 1: Regional pie chart
ws_regional = wb['Regional Summary']
chart_regional = PieChart()
chart_regional.title = 'Sales by Region'
data_ref = Reference(ws_regional, min_col=2, min_row=1, max_row=5)
labels_ref = Reference(ws_regional, min_col=1, min_row=2, max_row=5)
chart_regional.add_data(data_ref, titles_from_data=True)
chart_regional.set_categories(labels_ref)
ws_regional.add_chart(chart_regional, 'D2')

# Chart 2: Product bar chart
ws_product = wb['Product Summary']
chart_product = BarChart()
chart_product.type = 'col'
chart_product.title = 'Sales by Product'
data_ref = Reference(ws_product, min_col=2, min_row=1, max_row=5)
labels_ref = Reference(ws_product, min_col=1, min_row=2, max_row=5)
chart_product.add_data(data_ref, titles_from_data=True)
chart_product.set_categories(labels_ref)
ws_product.add_chart(chart_product, 'D2')

wb.save(multi_chart_file)
print(f"✓ Created: {multi_chart_file.name}")
print()

# ==============================================================================
# 10. DATE-BASED FORMULAS
# ==============================================================================

print("10. DATE-BASED FORMULAS")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Date Calculations"

# Headers
ws['A1'] = 'Employee'
ws['B1'] = 'Hire Date'
ws['C1'] = 'Today'
ws['D1'] = 'Days Employed'
ws['E1'] = 'Years Employed'

# Sample employees
employees = [
    ('Alice', '2020-01-15'),
    ('Bob', '2019-06-20'),
    ('Charlie', '2021-03-10'),
    ('Diana', '2018-11-05')
]

for row, (name, hire_date) in enumerate(employees, 2):
    ws[f'A{row}'] = name
    ws[f'B{row}'] = hire_date
    ws[f'B{row}'].number_format = 'YYYY-MM-DD'
    
    # TODAY formula
    ws[f'C{row}'] = '=TODAY()'
    ws[f'C{row}'].number_format = 'YYYY-MM-DD'
    
    # Days employed
    ws[f'D{row}'] = f'=C{row}-B{row}'
    
    # Years employed (approximate)
    ws[f'E{row}'] = f'=D{row}/365'
    ws[f'E{row}'].number_format = '0.0'

date_formulas_file = output_dir / '10_date_formulas.xlsx'
wb.save(date_formulas_file)
print(f"✓ Created: {date_formulas_file.name}")
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 80)
print("SESSION 6 COMPLETE!")
print("=" * 80)
print()
print("What you learned:")
print("✓ Write basic and advanced Excel formulas")
print("✓ Create column/bar charts")
print("✓ Create line charts for trends")
print("✓ Create pie charts")
print("✓ Create scatter plots")
print("✓ Use xlsxwriter for charts")
print("✓ Build comprehensive dashboards")
print("✓ Create multi-chart reports")
print("✓ Work with date-based formulas")
print("✓ Combine formulas and charts effectively")
print()
print("Files created:")
for file in sorted(output_dir.glob('*.xlsx')):
    print(f"  - {file.name}")
print()
print("Next: Session 7 - Merging and Combining Excel Files")
print("=" * 80)

