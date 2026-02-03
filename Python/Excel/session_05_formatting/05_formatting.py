"""
Session 5: Excel Formatting and Styling - Comprehensive Examples
================================================================

This script demonstrates comprehensive Excel formatting techniques
using openpyxl and xlsxwriter.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import xlsxwriter
from pathlib import Path
from datetime import datetime

print("=" * 80)
print("Session 5: Excel Formatting and Styling")
print("=" * 80)
print()

# Create output directory
output_dir = Path(__file__).parent / 'sample_files'
output_dir.mkdir(exist_ok=True)

# ==============================================================================
# 1. BASIC FONT STYLING
# ==============================================================================

print("1. BASIC FONT STYLING")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Font Styles"

# Bold
ws['A1'] = 'Bold Text'
ws['A1'].font = Font(bold=True)

# Italic
ws['A2'] = 'Italic Text'
ws['A2'].font = Font(italic=True)

# Bold + Italic
ws['A3'] = 'Bold and Italic'
ws['A3'].font = Font(bold=True, italic=True)

# Size and color
ws['B1'] = 'Large Text'
ws['B1'].font = Font(size=16)

ws['B2'] = 'Red Text'
ws['B2'].font = Font(color='FF0000')

ws['B3'] = 'Blue Bold Large'
ws['B3'].font = Font(bold=True, size=14, color='0000FF')

# Font families
ws['C1'] = 'Arial'
ws['C1'].font = Font(name='Arial', size=12)

ws['C2'] = 'Calibri'
ws['C2'].font = Font(name='Calibri', size=12)

ws['C3'] = 'Times New Roman'
ws['C3'].font = Font(name='Times New Roman', size=12)

font_file = output_dir / '01_font_styling.xlsx'
wb.save(font_file)
print(f"✓ Created: {font_file}")
print()

# ==============================================================================
# 2. CELL BACKGROUND COLORS
# ==============================================================================

print("2. CELL BACKGROUND COLORS")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Colors"

# Light colors
ws['A1'] = 'Light Blue'
ws['A1'].fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

ws['A2'] = 'Light Green'
ws['A2'].fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

ws['A3'] = 'Light Yellow'
ws['A3'].fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')

# Corporate/Professional colors
ws['B1'] = 'Primary Blue'
ws['B1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws['B1'].font = Font(color='FFFFFF', bold=True)

ws['B2'] = 'Dark Gray'
ws['B2'].fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
ws['B2'].font = Font(color='FFFFFF')

ws['B3'] = 'Dark Green'
ws['B3'].fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
ws['B3'].font = Font(color='FFFFFF', bold=True)

colors_file = output_dir / '02_cell_colors.xlsx'
wb.save(colors_file)
print(f"✓ Created: {colors_file}")
print()

# ==============================================================================
# 3. BORDERS
# ==============================================================================

print("3. BORDERS")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Borders"

# Thin border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ws['A1'] = 'Thin Border'
ws['A1'].border = thin_border

# Thick border
thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

ws['B1'] = 'Thick Border'
ws['B1'].border = thick_border

# Custom border (different sides)
custom_border = Border(
    top=Side(style='thick', color='FF0000'),
    bottom=Side(style='double'),
    left=Side(style='thin'),
    right=Side(style='thin')
)

ws['C1'] = 'Custom Border'
ws['C1'].border = custom_border

borders_file = output_dir / '03_borders.xlsx'
wb.save(borders_file)
print(f"✓ Created: {borders_file}")
print()

# ==============================================================================
# 4. ALIGNMENT
# ==============================================================================

print("4. ALIGNMENT")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Alignment"

# Horizontal alignment
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15

ws['A1'] = 'Left'
ws['A1'].alignment = Alignment(horizontal='left')

ws['B1'] = 'Center'
ws['B1'].alignment = Alignment(horizontal='center')

ws['C1'] = 'Right'
ws['C1'].alignment = Alignment(horizontal='right')

# Vertical alignment
ws.row_dimensions[3].height = 30

ws['A3'] = 'Top'
ws['A3'].alignment = Alignment(vertical='top')

ws['B3'] = 'Center'
ws['B3'].alignment = Alignment(vertical='center')

ws['C3'] = 'Bottom'
ws['C3'].alignment = Alignment(vertical='bottom')

# Text wrap
ws['A5'] = 'This is a long text that will wrap to multiple lines'
ws['A5'].alignment = Alignment(wrap_text=True)

alignment_file = output_dir / '04_alignment.xlsx'
wb.save(alignment_file)
print(f"✓ Created: {alignment_file}")
print()

# ==============================================================================
# 5. NUMBER FORMATTING
# ==============================================================================

print("5. NUMBER FORMATTING")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Number Formats"

# Currency
ws['A1'] = 'Currency'
ws['B1'] = 1234.56
ws['B1'].number_format = '$#,##0.00'

# Percentage
ws['A2'] = 'Percentage'
ws['B2'] = 0.856
ws['B2'].number_format = '0.00%'

# Date formats
ws['A3'] = 'Date Short'
ws['B3'] = datetime(2024, 3, 15)
ws['B3'].number_format = 'MM/DD/YYYY'

ws['A4'] = 'Date Long'
ws['B4'] = datetime(2024, 3, 15)
ws['B4'].number_format = 'MMMM DD, YYYY'

# Thousands separator
ws['A5'] = 'Thousands'
ws['B5'] = 1234567.89
ws['B5'].number_format = '#,##0.00'

number_format_file = output_dir / '05_number_formats.xlsx'
wb.save(number_format_file)
print(f"✓ Created: {number_format_file}")
print()

# ==============================================================================
# 6. CONDITIONAL FORMATTING
# ==============================================================================

print("6. CONDITIONAL FORMATTING")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Conditional"

# Add data
ws['A1'] = 'Value'
for i in range(2, 12):
    ws[f'A{i}'] = i * 10

# Highlight > 50 (red)
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_font = Font(color='9C0006')
rule1 = CellIsRule(operator='greaterThan', formula=['50'], fill=red_fill, font=red_font)
ws.conditional_formatting.add('A2:A11', rule1)

# Highlight < 30 (green)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_font = Font(color='006100')
rule2 = CellIsRule(operator='lessThan', formula=['30'], fill=green_fill, font=green_font)
ws.conditional_formatting.add('A2:A11', rule2)

conditional_file = output_dir / '06_conditional_formatting.xlsx'
wb.save(conditional_file)
print(f"✓ Created: {conditional_file}")
print()

# ==============================================================================
# 7. COLOR SCALES AND DATA BARS
# ==============================================================================

print("7. COLOR SCALES AND DATA BARS")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Visual Formatting"

# Color scale column
ws['A1'] = 'Color Scale'
for i in range(2, 22):
    ws[f'A{i}'] = i * 5

color_scale = ColorScaleRule(
    start_type='min', start_color='FF0000',
    mid_type='percentile', mid_value=50, mid_color='FFFF00',
    end_type='max', end_color='00FF00'
)
ws.conditional_formatting.add('A2:A21', color_scale)

# Data bars column
ws['B1'] = 'Data Bars'
revenues = [45000, 52000, 48000, 61000, 55000, 58000, 63000, 57000]
for i, revenue in enumerate(revenues, 2):
    ws[f'B{i}'] = revenue

data_bar = DataBarRule(start_type='min', end_type='max', color='4472C4')
ws.conditional_formatting.add('B2:B9', data_bar)

visual_file = output_dir / '07_visual_formatting.xlsx'
wb.save(visual_file)
print(f"✓ Created: {visual_file}")
print()

# ==============================================================================
# 8. XLSXWRITER FORMATTING
# ==============================================================================

print("8. XLSXWRITER FORMATTING")
print("-" * 80)

xlsxwriter_file = output_dir / '08_xlsxwriter_formatted.xlsx'
workbook = xlsxwriter.Workbook(xlsxwriter_file)
worksheet = workbook.add_worksheet('Report')

# Define formats
title_format = workbook.add_format({
    'bold': True,
    'font_size': 16,
    'align': 'center',
    'valign': 'vcenter',
    'fg_color': '#4472C4',
    'font_color': 'white'
})

header_format = workbook.add_format({
    'bold': True,
    'font_size': 11,
    'align': 'center',
    'valign': 'vcenter',
    'fg_color': '#D9E1F2',
    'border': 1
})

currency_format = workbook.add_format({
    'num_format': '$#,##0.00',
    'align': 'right',
    'border': 1
})

percent_format = workbook.add_format({
    'num_format': '0.0%',
    'align': 'right',
    'border': 1
})

text_format = workbook.add_format({
    'align': 'left',
    'border': 1
})

# Write content
worksheet.merge_range('A1:E1', 'Q1 2024 Sales Report', title_format)
worksheet.set_row(0, 30)

headers = ['Region', 'Sales', 'Target', 'Achievement', 'Status']
for col, header in enumerate(headers):
    worksheet.write(1, col, header, header_format)

data = [
    ['North', 125000, 100000, 1.25, 'Above Target'],
    ['South', 98000, 110000, 0.891, 'Below Target'],
    ['East', 142000, 130000, 1.092, 'Above Target'],
    ['West', 115000, 105000, 1.095, 'Above Target']
]

for row_idx, row_data in enumerate(data, 2):
    worksheet.write(row_idx, 0, row_data[0], text_format)
    worksheet.write(row_idx, 1, row_data[1], currency_format)
    worksheet.write(row_idx, 2, row_data[2], currency_format)
    worksheet.write(row_idx, 3, row_data[3], percent_format)
    worksheet.write(row_idx, 4, row_data[4], text_format)

worksheet.set_column('A:A', 15)
worksheet.set_column('B:C', 12)
worksheet.set_column('D:D', 12)
worksheet.set_column('E:E', 15)

workbook.close()
print(f"✓ Created: {xlsxwriter_file}")
print()

# ==============================================================================
# 9. PANDAS + OPENPYXL COMBINED
# ==============================================================================

print("9. PANDAS + OPENPYXL COMBINED")
print("-" * 80)

# Create data with pandas
df = pd.DataFrame({
    'Product': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Webcam'],
    'Q1': [45, 120, 85, 32, 67],
    'Q2': [52, 135, 90, 38, 71],
    'Q3': [48, 128, 88, 35, 69],
    'Q4': [56, 142, 95, 41, 75]
})

df['Total'] = df[['Q1', 'Q2', 'Q3', 'Q4']].sum(axis=1)
df['Average'] = df[['Q1', 'Q2', 'Q3', 'Q4']].mean(axis=1).round(1)

# Write with pandas
combined_file = output_dir / '09_combined_formatting.xlsx'
df.to_excel(combined_file, index=False, sheet_name='Sales')

# Format with openpyxl
wb = load_workbook(combined_file)
ws = wb['Sales']

# Define styles
header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format headers
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Format data
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = border
        if cell.column > 1:
            cell.alignment = Alignment(horizontal='right')

# Highlight totals
total_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=6).fill = total_fill
    ws.cell(row=row, column=6).font = Font(bold=True)

# Auto-fit columns
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max_length + 3

wb.save(combined_file)
print(f"✓ Created: {combined_file}")
print()

# ==============================================================================
# 10. PROFESSIONAL REPORT TEMPLATE
# ==============================================================================

print("10. PROFESSIONAL REPORT TEMPLATE")
print("-" * 80)

# Create comprehensive report
df_report = pd.DataFrame({
    'Department': ['Sales', 'Marketing', 'IT', 'HR', 'Operations'],
    'Budget': [500000, 200000, 350000, 150000, 400000],
    'Actual': [480000, 195000, 340000, 148000, 395000],
    'Employees': [45, 12, 28, 8, 35]
})

df_report['Variance'] = df_report['Budget'] - df_report['Actual']
df_report['Variance_%'] = ((df_report['Variance'] / df_report['Budget']) * 100).round(1)

# Write base file
professional_file = output_dir / '10_professional_report.xlsx'
df_report.to_excel(professional_file, index=False, sheet_name='Budget Report')

# Format extensively
wb = load_workbook(professional_file)
ws = wb['Budget Report']

# Insert company header rows
ws.insert_rows(1, 3)

ws['A1'] = 'COMPANY NAME'
ws['A1'].font = Font(bold=True, size=18, color='003366')
ws.merge_cells('A1:F1')

ws['A2'] = f'Budget Report - {datetime.now().strftime("%B %Y")}'
ws['A2'].font = Font(size=12, color='666666')
ws.merge_cells('A2:F2')

# Format header row (now row 4)
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

for cell in ws[4]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Format data rows
for row in range(5, ws.max_row + 1):
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        
        # Currency format
        if col in [2, 3, 5]:
            cell.number_format = '$#,##0'
        # Percentage format
        if col == 6:
            cell.number_format = '0.0%'

# Auto-fit columns
for idx, column in enumerate(ws.columns, start=1):
    max_length = 0
    column_letter = get_column_letter(idx)
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max(max_length + 3, 12)

wb.save(professional_file)
print(f"✓ Created: {professional_file}")
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 80)
print("SESSION 5 COMPLETE!")
print("=" * 80)
print()
print("What you learned:")
print("✓ Apply fonts, colors, and borders")
print("✓ Format numbers (currency, percentage, dates)")
print("✓ Adjust column widths and row heights")
print("✓ Use conditional formatting")
print("✓ Create color scales and data bars")
print("✓ Format with xlsxwriter")
print("✓ Combine pandas with openpyxl formatting")
print("✓ Create professional report templates")
print()
print("Files created:")
for file in sorted(output_dir.glob('*.xlsx')):
    print(f"  - {file.name}")
print()
print("Next: Session 6 - Working with Formulas and Charts")
print("=" * 80)

