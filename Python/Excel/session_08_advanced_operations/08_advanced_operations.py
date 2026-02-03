"""
Session 8: Advanced Data Operations - Comprehensive Examples
============================================================

This script demonstrates advanced Excel operations including named ranges,
data validation, worksheet protection, Excel tables, and more.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.comments import Comment
from openpyxl.styles import Protection, Font, PatternFill
from pathlib import Path
from datetime import datetime, date

print("=" * 80)
print("Session 8: Advanced Data Operations")
print("=" * 80)
print()

# Create output directory
output_dir = Path(__file__).parent / 'sample_files'
output_dir.mkdir(exist_ok=True)

# ==============================================================================
# 1. NAMED RANGES
# ==============================================================================

print("1. NAMED RANGES")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Sales Data"

# Add data
data = [
    ['Product', 'Price', 'Quantity'],
    ['Laptop', 1000, 5],
    ['Mouse', 25, 50],
    ['Keyboard', 75, 20]
]

for row in data:
    ws.append(row)

# Create named range for prices
defined_name = DefinedName('ProductPrices', attr_text=f"'{ws.title}'!$B$2:$B$4")
wb.defined_names['ProductPrices'] = defined_name

# Create named range for quantities
defined_name2 = DefinedName('ProductQuantities', attr_text=f"'{ws.title}'!$C$2:$C$4")
wb.defined_names['ProductQuantities'] = defined_name2

# Use named ranges in formulas
ws['E2'] = 'Total Value:'
ws['E3'] = '=SUMPRODUCT(ProductPrices, ProductQuantities)'

named_ranges_file = output_dir / '01_named_ranges.xlsx'
wb.save(named_ranges_file)
print(f"✓ Created: {named_ranges_file.name}")
print(f"  Named ranges: ProductPrices, ProductQuantities")
print()

# ==============================================================================
# 2. DATA VALIDATION - DROPDOWN LISTS
# ==============================================================================

print("2. DATA VALIDATION - DROPDOWN LISTS")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add headers
ws['A1'] = 'Department'
ws['B1'] = 'Employee'
ws['C1'] = 'Status'

# Create data validation for departments
dv_dept = DataValidation(
    type="list",
    formula1='"Sales,IT,HR,Marketing,Operations"',
    allow_blank=False
)
dv_dept.prompt = "Please select a department"
dv_dept.promptTitle = "Department Selection"
dv_dept.error = "Invalid department selected"
dv_dept.errorTitle = "Invalid Selection"

ws.add_data_validation(dv_dept)
dv_dept.add('A2:A100')

# Create validation for status
dv_status = DataValidation(
    type="list",
    formula1='"Active,Inactive,On Leave"'
)
dv_status.prompt = "Select employment status"
ws.add_data_validation(dv_status)
dv_status.add('C2:C100')

validation_file = output_dir / '02_data_validation.xlsx'
wb.save(validation_file)
print(f"✓ Created: {validation_file.name}")
print(f"  Dropdown validation added to columns A and C")
print()

# ==============================================================================
# 3. NUMERIC AND DATE VALIDATION
# ==============================================================================

print("3. NUMERIC AND DATE VALIDATION")
print("-" * 80)

wb = Workbook()
ws = wb.active

ws['A1'] = 'Age'
ws['B1'] = 'Salary'
ws['C1'] = 'Start Date'

# Age validation (18-65)
dv_age = DataValidation(
    type="whole",
    operator="between",
    formula1=18,
    formula2=65,
    allow_blank=False
)
dv_age.prompt = "Enter age between 18 and 65"
dv_age.error = "Age must be between 18 and 65"
ws.add_data_validation(dv_age)
dv_age.add('A2:A100')

# Salary validation (> 30000)
dv_salary = DataValidation(
    type="decimal",
    operator="greaterThan",
    formula1=30000
)
dv_salary.prompt = "Enter salary greater than $30,000"
dv_salary.error = "Salary must be greater than $30,000"
ws.add_data_validation(dv_salary)
dv_salary.add('B2:B100')

# Date validation (future dates only)
dv_date = DataValidation(
    type="date",
    operator="greaterThan",
    formula1=date.today()
)
dv_date.prompt = "Enter a future date"
dv_date.error = "Date must be in the future"
ws.add_data_validation(dv_date)
dv_date.add('C2:C100')

numeric_validation_file = output_dir / '03_numeric_validation.xlsx'
wb.save(numeric_validation_file)
print(f"✓ Created: {numeric_validation_file.name}")
print(f"  Age, salary, and date validation added")
print()

# ==============================================================================
# 4. WORKSHEET PROTECTION
# ==============================================================================

print("4. WORKSHEET PROTECTION")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add data
ws['A1'] = 'ID'
ws['B1'] = 'Name'
ws['C1'] = 'Editable Field'

ws['A2'] = 1
ws['B2'] = 'Alice'
ws['C2'] = 'You can edit this'

# Unlock specific cells (C column)
for row in range(2, 100):
    ws[f'C{row}'].protection = Protection(locked=False)

# Protect worksheet
ws.protection.sheet = True
ws.protection.password = 'secret123'

protected_file = output_dir / '04_protected_worksheet.xlsx'
wb.save(protected_file)
print(f"✓ Created: {protected_file.name}")
print(f"  Worksheet protected (password: secret123)")
print(f"  Column C is editable")
print()

# ==============================================================================
# 5. EXCEL TABLES
# ==============================================================================

print("5. EXCEL TABLES")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add data
data = [
    ['Product', 'Q1', 'Q2', 'Q3', 'Q4'],
    ['Laptop', 100, 120, 110, 130],
    ['Mouse', 500, 550, 520, 580],
    ['Keyboard', 300, 320, 310, 340],
    ['Monitor', 150, 170, 160, 180]
]

for row in data:
    ws.append(row)

# Create Excel table
tab = Table(displayName="SalesTable", ref="A1:E5")

style = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style

ws.add_table(tab)

excel_table_file = output_dir / '05_excel_table.xlsx'
wb.save(excel_table_file)
print(f"✓ Created: {excel_table_file.name}")
print(f"  Excel table 'SalesTable' created with styling")
print()

# ==============================================================================
# 6. FREEZE PANES
# ==============================================================================

print("6. FREEZE PANES")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add headers
headers = ['ID', 'Name', 'Department', 'Salary', 'Bonus', 'Total']
ws.append(headers)

# Add many rows
for i in range(1, 101):
    ws.append([i, f'Employee{i}', 'Sales', 50000, 5000, 55000])

# Freeze top row
ws.freeze_panes = 'A2'

freeze_file = output_dir / '06_freeze_panes.xlsx'
wb.save(freeze_file)
print(f"✓ Created: {freeze_file.name}")
print(f"  Top row frozen (100 data rows)")
print()

# ==============================================================================
# 7. AUTOFILTER
# ==============================================================================

print("7. AUTOFILTER")
print("-" * 80)

# Create sample data with pandas
df = pd.DataFrame({
    'Name': ['Alice', 'Bob', 'Charlie', 'Diana', 'Eve'] * 4,
    'Department': ['Sales', 'IT', 'HR', 'Marketing', 'Operations'] * 4,
    'Salary': np.random.randint(40000, 100000, 20),
    'Status': ['Active'] * 15 + ['Inactive'] * 5
})

autofilter_file = output_dir / '07_autofilter.xlsx'
df.to_excel(autofilter_file, index=False)

# Add autofilter
wb = load_workbook(autofilter_file)
ws = wb.active
ws.auto_filter.ref = ws.dimensions

wb.save(autofilter_file)
print(f"✓ Created: {autofilter_file.name}")
print(f"  AutoFilter enabled on all columns")
print()

# ==============================================================================
# 8. COMMENTS AND NOTES
# ==============================================================================

print("8. COMMENTS AND NOTES")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add data with comments
ws['A1'] = 'Important Data'
ws['A1'].comment = Comment(
    'This value was verified on 2024-03-15.\nDo not change without approval.',
    'John Doe'
)

ws['B1'] = 'Target'
ws['B1'].comment = Comment('Q1 target value', 'Manager')

ws['A2'] = 1000
ws['A2'].comment = Comment('Verified value', 'Auditor')

comments_file = output_dir / '08_with_comments.xlsx'
wb.save(comments_file)
print(f"✓ Created: {comments_file.name}")
print(f"  3 cells have comments attached")
print()

# ==============================================================================
# 9. HIDDEN ROWS AND COLUMNS
# ==============================================================================

print("9. HIDDEN ROWS AND COLUMNS")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add data
for i in range(1, 11):
    ws.append([f'Row {i}', i * 100, i * 200, i * 300])

# Hide specific rows (intermediate calculations)
ws.row_dimensions[3].hidden = True
ws.row_dimensions[5].hidden = True
ws.row_dimensions[7].hidden = True

# Hide column C (intermediate data)
ws.column_dimensions['C'].hidden = True

hidden_file = output_dir / '09_hidden_rows_cols.xlsx'
wb.save(hidden_file)
print(f"✓ Created: {hidden_file.name}")
print(f"  Rows 3, 5, 7 and column C are hidden")
print()

# ==============================================================================
# 10. GROUPING (OUTLINING)
# ==============================================================================

print("10. GROUPING (OUTLINING)")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add hierarchical data
ws.append(['Category', 'Subcategory', 'Value'])
ws.append(['Sales', 'Q1', 100])
ws.append(['Sales', 'Q2', 120])
ws.append(['Sales', 'Q3', 110])
ws.append(['Sales', 'Q4', 130])
ws.append(['Marketing', 'Q1', 50])
ws.append(['Marketing', 'Q2', 55])
ws.append(['Marketing', 'Q3', 52])
ws.append(['Marketing', 'Q4', 58])

# Group rows
ws.row_dimensions.group(2, 5, outline_level=1)  # Sales quarters
ws.row_dimensions.group(6, 9, outline_level=1)  # Marketing quarters

grouped_file = output_dir / '10_grouped_rows.xlsx'
wb.save(grouped_file)
print(f"✓ Created: {grouped_file.name}")
print(f"  Rows grouped by category")
print()

# ==============================================================================
# 11. DOCUMENT PROPERTIES AND METADATA
# ==============================================================================

print("11. DOCUMENT PROPERTIES AND METADATA")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Set document properties
wb.properties.title = 'Monthly Sales Report'
wb.properties.subject = 'Sales Data'
wb.properties.creator = 'Python Automation'
wb.properties.keywords = 'sales, monthly, report'
wb.properties.description = 'Comprehensive sales analysis for March 2024'
wb.properties.category = 'Financial Reports'
wb.properties.company = 'Acme Corporation'
wb.properties.created = datetime(2024, 3, 1)
wb.properties.modified = datetime.now()

# Add some data
ws['A1'] = 'Sales Report'
ws['A2'] = 'Generated by Python'

metadata_file = output_dir / '11_with_metadata.xlsx'
wb.save(metadata_file)
print(f"✓ Created: {metadata_file.name}")
print(f"  Document properties set:")
print(f"    Title: {wb.properties.title}")
print(f"    Company: {wb.properties.company}")
print(f"    Creator: {wb.properties.creator}")
print()

# ==============================================================================
# 12. COMPREHENSIVE EXAMPLE - EMPLOYEE DATABASE
# ==============================================================================

print("12. COMPREHENSIVE EXAMPLE - EMPLOYEE DATABASE")
print("-" * 80)

wb = Workbook()
ws = wb.active
ws.title = "Employees"

# Setup headers
headers = ['ID', 'Name', 'Department', 'Salary', 'Status', 'Start Date']
ws.append(headers)

# Format headers
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Add sample data
sample_employees = [
    [1, 'Alice Smith', 'Sales', 60000, 'Active', date(2020, 1, 15)],
    [2, 'Bob Jones', 'IT', 75000, 'Active', date(2019, 6, 20)],
    [3, 'Charlie Brown', 'HR', 58000, 'Active', date(2021, 3, 10)],
    [4, 'Diana Lee', 'Sales', 62000, 'Active', date(2020, 8, 5)],
    [5, 'Eve Wilson', 'IT', 72000, 'Active', date(2018, 11, 22)]
]

for emp in sample_employees:
    ws.append(emp)

# Add data validation for department
dv_dept = DataValidation(
    type="list",
    formula1='"Sales,IT,HR,Marketing,Operations"'
)
ws.add_data_validation(dv_dept)
dv_dept.add('C2:C100')

# Add validation for status
dv_status = DataValidation(
    type="list",
    formula1='"Active,Inactive,On Leave"'
)
ws.add_data_validation(dv_status)
dv_status.add('E2:E100')

# Protect ID column
for row in range(2, 100):
    ws[f'A{row}'].protection = Protection(locked=True)

# Create Excel table
tab = Table(displayName="EmployeeTable", ref="A1:F6")
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showRowStripes=True
)
tab.tableStyleInfo = style
ws.add_table(tab)

# Add autofilter
ws.auto_filter.ref = "A1:F6"

# Freeze header row
ws.freeze_panes = 'A2'

# Add comments
ws['A1'].comment = Comment('Employee ID - Auto-generated', 'System')
ws['D1'].comment = Comment('Salary in USD', 'HR Department')

comprehensive_file = output_dir / '12_comprehensive_employee_db.xlsx'
wb.save(comprehensive_file)
print(f"✓ Created: {comprehensive_file.name}")
print(f"  Features included:")
print(f"    - Excel table with styling")
print(f"    - Data validation on Department and Status")
print(f"    - Protected ID column")
print(f"    - AutoFilter enabled")
print(f"    - Frozen header row")
print(f"    - Comments on key fields")
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 80)
print("SESSION 8 COMPLETE!")
print("=" * 80)
print()
print("What you learned:")
print("✓ Create and use named ranges")
print("✓ Add data validation (dropdowns, numeric, dates)")
print("✓ Protect worksheets and specific cells")
print("✓ Create Excel tables with styling")
print("✓ Freeze panes for better navigation")
print("✓ Add autofilters")
print("✓ Add comments to cells")
print("✓ Hide rows and columns")
print("✓ Group rows (outlining)")
print("✓ Set document properties")
print("✓ Build comprehensive database with all features")
print()
print("Files created:")
for file in sorted(output_dir.glob('*.xlsx')):
    print(f"  - {file.name}")
print()
print("Next: Session 9 - Working with Large Excel Files")
print("=" * 80)

