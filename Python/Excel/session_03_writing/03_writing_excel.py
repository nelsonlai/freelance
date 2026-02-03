"""
Session 3: Writing Excel Files - Comprehensive Examples
=======================================================

This script demonstrates comprehensive techniques for writing Excel files
with various formats, structures, and requirements.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
from pathlib import Path
from datetime import datetime

print("=" * 80)
print("Session 3: Writing Excel Files - Comprehensive Techniques")
print("=" * 80)
print()

# Create output directory
output_dir = Path(__file__).parent / 'sample_files'
output_dir.mkdir(exist_ok=True)

# ==============================================================================
# 1. BASIC WRITING WITH PANDAS
# ==============================================================================

print("1. BASIC WRITING WITH PANDAS")
print("-" * 80)

# Create sample data
df_products = pd.DataFrame({
    'Product_ID': range(101, 111),
    'Product_Name': [f'Product {i}' for i in range(1, 11)],
    'Category': np.random.choice(['Electronics', 'Clothing', 'Food'], 10),
    'Price': np.random.uniform(10, 500, 10).round(2),
    'Stock': np.random.randint(10, 200, 10)
})

print("Sample DataFrame:")
print(df_products)
print()

# Write to Excel
basic_file = output_dir / 'basic_output.xlsx'
df_products.to_excel(basic_file, index=False, sheet_name='Products')
print(f"✓ Created: {basic_file}")
print()

# Write with index
with_index_file = output_dir / 'with_index.xlsx'
df_products.to_excel(with_index_file, index=True, sheet_name='Products')
print(f"✓ Created: {with_index_file}")
print()

# Write without headers
no_header_file = output_dir / 'no_headers.xlsx'
df_products.to_excel(no_header_file, index=False, header=False, sheet_name='Data')
print(f"✓ Created: {no_header_file}")
print()

# ==============================================================================
# 2. WRITING MULTIPLE SHEETS
# ==============================================================================

print("2. WRITING MULTIPLE SHEETS")
print("-" * 80)

# Create multiple DataFrames
df_sales = pd.DataFrame({
    'Month': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
    'Revenue': [45000, 52000, 48000, 55000, 61000, 58000],
    'Expenses': [30000, 33000, 31000, 35000, 38000, 36000]
})
df_sales['Profit'] = df_sales['Revenue'] - df_sales['Expenses']

df_employees = pd.DataFrame({
    'Employee_ID': range(1, 6),
    'Name': ['Alice', 'Bob', 'Charlie', 'Diana', 'Eve'],
    'Department': ['Sales', 'IT', 'HR', 'Sales', 'IT'],
    'Salary': [60000, 75000, 58000, 65000, 72000]
})

df_customers = pd.DataFrame({
    'Customer_ID': range(1001, 1006),
    'Company': [f'Company {chr(65+i)}' for i in range(5)],
    'Contact': [f'contact{i}@email.com' for i in range(1, 6)],
    'Total_Orders': np.random.randint(5, 50, 5)
})

# Write all sheets using ExcelWriter
multi_sheet_file = output_dir / 'multi_sheet_report.xlsx'
with pd.ExcelWriter(multi_sheet_file, engine='openpyxl') as writer:
    df_sales.to_excel(writer, sheet_name='Monthly_Sales', index=False)
    df_employees.to_excel(writer, sheet_name='Employees', index=False)
    df_customers.to_excel(writer, sheet_name='Customers', index=False)

print(f"✓ Created multi-sheet workbook: {multi_sheet_file}")
print(f"  Sheets: Monthly_Sales, Employees, Customers")
print()

# ==============================================================================
# 3. APPENDING TO EXISTING FILES
# ==============================================================================

print("3. APPENDING TO EXISTING FILES")
print("-" * 80)

# Create initial file
initial_data = pd.DataFrame({
    'Date': pd.date_range('2024-01-01', periods=5),
    'Value': np.random.randint(100, 500, 5)
})

append_file = output_dir / 'appendable.xlsx'
initial_data.to_excel(append_file, sheet_name='Data', index=False)
print(f"✓ Created initial file: {append_file}")
print()

# Append new sheet to existing file
new_data = pd.DataFrame({
    'Product': ['A', 'B', 'C'],
    'Quantity': [100, 150, 200]
})

with pd.ExcelWriter(append_file, mode='a', engine='openpyxl') as writer:
    new_data.to_excel(writer, sheet_name='New_Data', index=False)

print(f"✓ Appended new sheet to: {append_file}")
print()

# ==============================================================================
# 4. WRITING WITH OPENPYXL
# ==============================================================================

print("4. WRITING WITH OPENPYXL")
print("-" * 80)

# Create new workbook
wb = Workbook()
ws = wb.active
ws.title = "Sample Data"

# Write headers
headers = ['ID', 'Name', 'Score', 'Grade']
ws.append(headers)

# Write data rows
data = [
    [1, 'Alice', 95, 'A'],
    [2, 'Bob', 87, 'B'],
    [3, 'Charlie', 92, 'A'],
    [4, 'Diana', 78, 'C'],
    [5, 'Eve', 88, 'B']
]

for row in data:
    ws.append(row)

openpyxl_file = output_dir / 'openpyxl_basic.xlsx'
wb.save(openpyxl_file)
print(f"✓ Created with openpyxl: {openpyxl_file}")
print()

# ==============================================================================
# 5. WRITING WITH FORMATTING
# ==============================================================================

print("5. WRITING WITH FORMATTING (OPENPYXL)")
print("-" * 80)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Formatted Report"

# Write and format header
header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center')

headers = ['Product', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales', 'Total']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Write data with formatting
products = ['Product A', 'Product B', 'Product C', 'Product D']
for row_num, product in enumerate(products, 2):
    ws.cell(row=row_num, column=1, value=product)
    
    # Add quarterly sales
    q_sales = []
    for col in range(2, 6):
        value = np.random.randint(10000, 50000)
        ws.cell(row=row_num, column=col, value=value)
        q_sales.append(value)
    
    # Add total with formula
    ws.cell(row=row_num, column=6, value=f'=SUM(B{row_num}:E{row_num})')

# Adjust column widths
for col in range(1, 7):
    ws.column_dimensions[chr(64 + col)].width = 15

formatted_file = output_dir / 'formatted_report.xlsx'
wb.save(formatted_file)
print(f"✓ Created formatted file: {formatted_file}")
print()

# ==============================================================================
# 6. WRITING WITH XLSXWRITER
# ==============================================================================

print("6. WRITING WITH XLSXWRITER")
print("-" * 80)

xlsxwriter_file = output_dir / 'xlsxwriter_example.xlsx'
workbook = xlsxwriter.Workbook(xlsxwriter_file)
worksheet = workbook.add_worksheet('Sales Report')

# Create formats
header_format = workbook.add_format({
    'bold': True,
    'font_size': 12,
    'bg_color': '#4472C4',
    'font_color': 'white',
    'align': 'center',
    'valign': 'vcenter',
    'border': 1
})

currency_format = workbook.add_format({
    'num_format': '$#,##0.00',
    'align': 'right'
})

percent_format = workbook.add_format({
    'num_format': '0.0%',
    'align': 'right'
})

# Write headers
headers = ['Region', 'Sales', 'Target', 'Achievement']
for col, header in enumerate(headers):
    worksheet.write(0, col, header, header_format)

# Write data
regions_data = [
    ['North', 125000, 100000],
    ['South', 98000, 110000],
    ['East', 142000, 130000],
    ['West', 115000, 105000]
]

for row, (region, sales, target) in enumerate(regions_data, 1):
    worksheet.write(row, 0, region)
    worksheet.write(row, 1, sales, currency_format)
    worksheet.write(row, 2, target, currency_format)
    # Achievement percentage formula
    worksheet.write_formula(row, 3, f'=B{row+1}/C{row+1}', percent_format)

# Set column widths
worksheet.set_column('A:A', 15)
worksheet.set_column('B:C', 12)
worksheet.set_column('D:D', 12)

workbook.close()
print(f"✓ Created with xlsxwriter: {xlsxwriter_file}")
print()

# ==============================================================================
# 7. COMBINING PANDAS AND OPENPYXL
# ==============================================================================

print("7. COMBINING PANDAS AND OPENPYXL FOR ADVANCED FORMATTING")
print("-" * 80)

# Create data with pandas
summary_data = pd.DataFrame({
    'Department': ['Sales', 'Marketing', 'IT', 'HR', 'Operations'],
    'Employees': [45, 12, 28, 8, 35],
    'Budget': [500000, 200000, 350000, 150000, 400000],
    'Actual': [480000, 195000, 340000, 170000, 395000]
})
summary_data['Variance'] = summary_data['Budget'] - summary_data['Actual']
summary_data['Variance_Pct'] = (summary_data['Variance'] / summary_data['Budget'] * 100).round(2)

# Write with pandas
combined_file = output_dir / 'combined_approach.xlsx'
summary_data.to_excel(combined_file, sheet_name='Summary', index=False)

# Format with openpyxl
wb = load_workbook(combined_file)
ws = wb['Summary']

# Format header
for cell in ws[1]:
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Format currency columns
currency_cols = ['C', 'D', 'E']  # Budget, Actual, Variance
for col in currency_cols:
    for cell in ws[col][1:]:  # Skip header
        cell.number_format = '$#,##0'

# Format percentage column
for cell in ws['F'][1:]:  # Variance_Pct
    cell.number_format = '0.00%'

# Add conditional formatting for Variance column
from openpyxl.formatting.rule import CellIsRule

red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

ws.conditional_formatting.add('E2:E6', 
    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
ws.conditional_formatting.add('E2:E6', 
    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))

# Adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max_length + 3

wb.save(combined_file)
print(f"✓ Created with combined approach: {combined_file}")
print()

# ==============================================================================
# 8. WRITING DATAFRAME WITH CUSTOM START POSITION
# ==============================================================================

print("8. WRITING WITH CUSTOM START POSITION")
print("-" * 80)

wb = Workbook()
ws = wb.active

# Add title
ws['A1'] = 'Monthly Sales Report'
ws['A1'].font = Font(size=16, bold=True)
ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d")}'

# Write DataFrame starting from row 4
for r in dataframe_to_rows(df_sales, index=False, header=True):
    ws.append(r)

# Adjust placement (move data down)
custom_position_file = output_dir / 'custom_position.xlsx'
wb.save(custom_position_file)
print(f"✓ Created with custom positioning: {custom_position_file}")
print()

# ==============================================================================
# 9. PERFORMANCE COMPARISON
# ==============================================================================

print("9. PERFORMANCE COMPARISON")
print("-" * 80)

import time

# Create large dataset
large_df = pd.DataFrame({
    'Col_' + str(i): np.random.randn(10000) for i in range(10)
})

# Test pandas writing
large_file = output_dir / 'large_pandas.xlsx'
start = time.time()
large_df.to_excel(large_file, index=False)
pandas_time = time.time() - start
print(f"Pandas write time (10,000 rows): {pandas_time:.4f} seconds")

# Test openpyxl writing
large_file_openpyxl = output_dir / 'large_openpyxl.xlsx'
start = time.time()
wb = Workbook()
ws = wb.active
for row in dataframe_to_rows(large_df, index=False, header=True):
    ws.append(row)
wb.save(large_file_openpyxl)
openpyxl_time = time.time() - start
print(f"Openpyxl write time (10,000 rows): {openpyxl_time:.4f} seconds")

if pandas_time < openpyxl_time:
    print(f"✓ Pandas is {openpyxl_time/pandas_time:.2f}x faster for large datasets")
else:
    print(f"✓ Openpyxl is {pandas_time/openpyxl_time:.2f}x faster for large datasets")
print()

# ==============================================================================
# 10. ERROR HANDLING
# ==============================================================================

print("10. ERROR HANDLING")
print("-" * 80)

def safe_write_excel(df, file_path, **kwargs):
    """Safely write DataFrame to Excel with error handling"""
    try:
        file_path = Path(file_path)
        
        # Check if DataFrame is empty
        if df.empty:
            print(f"⚠ Warning: DataFrame is empty, not writing to {file_path}")
            return False
        
        # Check if directory exists
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Write file
        df.to_excel(file_path, **kwargs)
        print(f"✓ Successfully wrote {len(df)} rows to {file_path}")
        return True
        
    except PermissionError:
        print(f"✗ Error: Permission denied. File may be open: {file_path}")
        return False
    
    except Exception as e:
        print(f"✗ Error writing file: {type(e).__name__}: {e}")
        return False

# Test safe writing
test_df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
safe_file = output_dir / 'safe_write_test.xlsx'
safe_write_excel(test_df, safe_file, index=False)
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 80)
print("SESSION 3 COMPLETE!")
print("=" * 80)
print()
print("What you learned:")
print("✓ Basic writing with pandas")
print("✓ Creating multi-sheet workbooks")
print("✓ Appending to existing files")
print("✓ Writing with openpyxl for more control")
print("✓ Adding formatting and styles")
print("✓ Using xlsxwriter for advanced formatting")
print("✓ Combining pandas and openpyxl")
print("✓ Performance optimization")
print("✓ Error handling best practices")
print()
print("Files created:")
for file in sorted(output_dir.glob('*.xlsx')):
    print(f"  - {file.name}")
print()
print("Next: Session 4 - Data Manipulation with Pandas")
print("=" * 80)

