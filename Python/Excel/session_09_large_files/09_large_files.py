"""
Session 9: Working with Large Excel Files - Comprehensive Examples
==================================================================

This script demonstrates techniques for efficiently handling large Excel files
including chunking, streaming, and optimization strategies.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
import time
from pathlib import Path
import os

print("=" * 80)
print("Session 9: Working with Large Excel Files")
print("=" * 80)
print()

# Create output directory
output_dir = Path(__file__).parent / 'sample_files'
output_dir.mkdir(exist_ok=True)

# ==============================================================================
# 1. CREATE SAMPLE LARGE FILES
# ==============================================================================

print("1. CREATING SAMPLE LARGE FILES")
print("-" * 80)

def create_sample_file(filename, num_rows=10000):
    """Create sample file with specified number of rows"""
    np.random.seed(42)
    
    df = pd.DataFrame({
        'ID': range(1, num_rows + 1),
        'Product': np.random.choice(['Laptop', 'Mouse', 'Keyboard', 'Monitor'], num_rows),
        'Region': np.random.choice(['North', 'South', 'East', 'West'], num_rows),
        'Sales': np.random.randint(100, 10000, num_rows),
        'Quantity': np.random.randint(1, 100, num_rows),
        'Date': pd.date_range('2024-01-01', periods=num_rows, freq='H')
    })
    
    df.to_excel(filename, index=False)
    file_size = Path(filename).stat().st_size / (1024 * 1024)  # MB
    return len(df), file_size

# Create small file
small_file = output_dir / 'small_file.xlsx'
rows, size = create_sample_file(small_file, 1000)
print(f"✓ Created {small_file.name}: {rows:,} rows, {size:.2f} MB")

# Create medium file
medium_file = output_dir / 'medium_file.xlsx'
rows, size = create_sample_file(medium_file, 5000)
print(f"✓ Created {medium_file.name}: {rows:,} rows, {size:.2f} MB")

# Create large file
large_file = output_dir / 'large_file.xlsx'
rows, size = create_sample_file(large_file, 10000)
print(f"✓ Created {large_file.name}: {rows:,} rows, {size:.2f} MB")

print()

# ==============================================================================
# 2. MEASURE FILE SIZE
# ==============================================================================

print("2. MEASURING FILE SIZE")
print("-" * 80)

def check_file_size(file_path):
    """Check file size and categorize"""
    path = Path(file_path)
    size_bytes = path.stat().st_size
    size_mb = size_bytes / (1024 * 1024)
    
    print(f"File: {path.name}")
    print(f"Size: {size_mb:.2f} MB")
    
    if size_mb < 10:
        category = "Small - normal processing OK"
    elif size_mb < 50:
        category = "Medium - consider chunking"
    elif size_mb < 200:
        category = "Large - use chunking/streaming"
    else:
        category = "Very large - use optimized strategies"
    
    print(f"Category: {category}")
    print()
    return size_mb

check_file_size(small_file)
check_file_size(medium_file)
check_file_size(large_file)

# ==============================================================================
# 3. READING WITH PERFORMANCE MEASUREMENT
# ==============================================================================

print("3. READING WITH PERFORMANCE MEASUREMENT")
print("-" * 80)

def timed_read(file_path, method="normal"):
    """Read file and measure time"""
    start_time = time.time()
    
    if method == "normal":
        df = pd.read_excel(file_path)
    elif method == "limited_cols":
        df = pd.read_excel(file_path, usecols=['ID', 'Product', 'Sales'])
    elif method == "limited_rows":
        df = pd.read_excel(file_path, nrows=1000)
    
    elapsed = time.time() - start_time
    
    return df, elapsed

# Test different reading methods
print(f"Reading {small_file.name}:")
df1, time1 = timed_read(small_file, "normal")
print(f"  Normal read: {time1:.2f} seconds, {len(df1):,} rows")

df2, time2 = timed_read(small_file, "limited_cols")
print(f"  Limited columns: {time2:.2f} seconds, {len(df2.columns)} cols")

df3, time3 = timed_read(small_file, "limited_rows")
print(f"  Limited rows: {time3:.2f} seconds, {len(df3):,} rows")
print()

# ==============================================================================
# 4. CHUNKED READING
# ==============================================================================

print("4. CHUNKED READING")
print("-" * 80)

def read_in_chunks(file_path, chunk_size=2000):
    """Read large file in chunks and process"""
    print(f"Reading {Path(file_path).name} in chunks of {chunk_size}")
    
    results = []
    skip_rows = 0
    chunk_num = 0
    
    while True:
        try:
            if skip_rows > 0:
                df_chunk = pd.read_excel(
                    file_path,
                    skiprows=range(1, skip_rows + 1),
                    nrows=chunk_size
                )
            else:
                df_chunk = pd.read_excel(file_path, nrows=chunk_size)
            
            if df_chunk.empty:
                break
            
            chunk_num += 1
            
            # Process chunk (example: calculate sum)
            chunk_result = df_chunk['Sales'].sum()
            results.append(chunk_result)
            
            print(f"  Chunk {chunk_num}: {len(df_chunk):,} rows, Sales sum: ${chunk_result:,}")
            
            skip_rows += chunk_size
            
            if len(df_chunk) < chunk_size:
                break
                
        except Exception as e:
            print(f"  Error: {e}")
            break
    
    total = sum(results)
    print(f"  Total from {chunk_num} chunks: ${total:,}")
    print()
    return total

# Read large file in chunks
total_sales = read_in_chunks(large_file, chunk_size=2500)

# ==============================================================================
# 5. STREAMING WITH OPENPYXL (READ-ONLY MODE)
# ==============================================================================

print("5. STREAMING WITH OPENPYXL (READ-ONLY MODE)")
print("-" * 80)

def stream_read(file_path):
    """Read file using openpyxl streaming (memory efficient)"""
    print(f"Streaming {Path(file_path).name} with openpyxl...")
    
    start_time = time.time()
    
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    
    row_count = 0
    total_sales = 0
    
    # Skip header
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_count += 1
        if row[3] is not None:  # Sales column
            total_sales += row[3]
        
        if row_count % 2000 == 0:
            print(f"  Processed {row_count:,} rows...")
    
    wb.close()
    
    elapsed = time.time() - start_time
    
    print(f"  Completed: {row_count:,} rows in {elapsed:.2f} seconds")
    print(f"  Total sales: ${total_sales:,}")
    print()
    
    return row_count, total_sales

# Stream read large file
rows, sales = stream_read(large_file)

# ==============================================================================
# 6. WRITE-ONLY MODE (MEMORY EFFICIENT WRITING)
# ==============================================================================

print("6. WRITE-ONLY MODE (MEMORY EFFICIENT WRITING)")
print("-" * 80)

def write_large_file(output_file, num_rows=10000):
    """Write large file using write-only mode"""
    print(f"Writing {num_rows:,} rows in write-only mode...")
    
    start_time = time.time()
    
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    
    # Write header
    ws.append(['ID', 'Value', 'Category', 'Amount'])
    
    # Write data
    for i in range(1, num_rows + 1):
        ws.append([
            i,
            np.random.randint(100, 1000),
            np.random.choice(['A', 'B', 'C']),
            np.random.uniform(10.0, 1000.0)
        ])
        
        if i % 2000 == 0:
            print(f"  Wrote {i:,} rows...")
    
    wb.save(output_file)
    
    elapsed = time.time() - start_time
    file_size = Path(output_file).stat().st_size / (1024 * 1024)
    
    print(f"  Completed in {elapsed:.2f} seconds")
    print(f"  File size: {file_size:.2f} MB")
    print()

write_only_file = output_dir / 'write_only_large.xlsx'
write_large_file(write_only_file, 10000)

# ==============================================================================
# 7. OPTIMIZING DATA TYPES
# ==============================================================================

print("7. OPTIMIZING DATA TYPES")
print("-" * 80)

# Read with default types
df_default = pd.read_excel(small_file)
memory_default = df_default.memory_usage(deep=True).sum() / (1024 * 1024)
print(f"Default types memory: {memory_default:.2f} MB")
print(df_default.dtypes)
print()

# Read and optimize types
df_optimized = pd.read_excel(small_file)
df_optimized['ID'] = df_optimized['ID'].astype('int32')
df_optimized['Product'] = df_optimized['Product'].astype('category')
df_optimized['Region'] = df_optimized['Region'].astype('category')
df_optimized['Sales'] = df_optimized['Sales'].astype('int32')
df_optimized['Quantity'] = df_optimized['Quantity'].astype('int16')

memory_optimized = df_optimized.memory_usage(deep=True).sum() / (1024 * 1024)
print(f"Optimized types memory: {memory_optimized:.2f} MB")
print(df_optimized.dtypes)
print(f"Memory saved: {memory_default - memory_optimized:.2f} MB ({((memory_default - memory_optimized) / memory_default * 100):.1f}%)")
print()

# ==============================================================================
# 8. READING ONLY NEEDED COLUMNS
# ==============================================================================

print("8. READING ONLY NEEDED COLUMNS")
print("-" * 80)

# Time full read
start = time.time()
df_full = pd.read_excel(small_file)
time_full = time.time() - start

# Time partial read
start = time.time()
df_partial = pd.read_excel(small_file, usecols=['Product', 'Sales', 'Region'])
time_partial = time.time() - start

print(f"Full read: {len(df_full.columns)} columns, {time_full:.3f} seconds")
print(f"Partial read: {len(df_partial.columns)} columns, {time_partial:.3f} seconds")
print(f"Time saved: {time_full - time_partial:.3f} seconds ({((time_full - time_partial) / time_full * 100):.1f}%)")
print()

# ==============================================================================
# 9. CSV CONVERSION FOR SPEED
# ==============================================================================

print("9. CSV CONVERSION FOR SPEED")
print("-" * 80)

# Excel read
start = time.time()
df_excel = pd.read_excel(small_file)
time_excel = time.time() - start

# Convert to CSV
csv_file = output_dir / 'converted.csv'
df_excel.to_csv(csv_file, index=False)

# CSV read
start = time.time()
df_csv = pd.read_csv(csv_file)
time_csv = time.time() - start

print(f"Excel read: {time_excel:.3f} seconds")
print(f"CSV read: {time_csv:.3f} seconds")
print(f"Speedup: {time_excel / time_csv:.1f}x faster")
print()

# ==============================================================================
# 10. BATCH PROCESSING MULTIPLE FILES
# ==============================================================================

print("10. BATCH PROCESSING MULTIPLE FILES")
print("-" * 80)

def process_multiple_files(directory, pattern='*.xlsx'):
    """Process all files in directory"""
    files = list(Path(directory).glob(pattern))
    
    print(f"Found {len(files)} files to process")
    
    results = []
    for file in files:
        if file.name.startswith('~$'):  # Skip temp files
            continue
        
        try:
            df = pd.read_excel(file)
            result = {
                'File': file.name,
                'Rows': len(df),
                'Columns': len(df.columns),
                'Size_MB': file.stat().st_size / (1024 * 1024)
            }
            results.append(result)
            print(f"  ✓ {file.name}: {len(df):,} rows")
        except Exception as e:
            print(f"  ✗ {file.name}: {e}")
    
    return pd.DataFrame(results)

summary = process_multiple_files(output_dir)
print("\nSummary:")
print(summary.to_string(index=False))
print()

# ==============================================================================
# 11. PRODUCTION-READY LARGE FILE PROCESSOR
# ==============================================================================

print("11. PRODUCTION-READY LARGE FILE PROCESSOR")
print("-" * 80)

def production_file_processor(input_file, output_file, chunk_size=5000):
    """
    Production-ready processor with error handling and progress
    """
    print(f"Processing: {Path(input_file).name}")
    
    try:
        # Check file size
        file_size_mb = Path(input_file).stat().st_size / (1024 * 1024)
        print(f"  File size: {file_size_mb:.2f} MB")
        
        # Choose strategy
        if file_size_mb < 10:
            # Normal processing
            print("  Strategy: Normal read")
            df = pd.read_excel(input_file)
            
            # Process (example)
            df['Total'] = df['Sales'] * df['Quantity']
            
            # Save
            df.to_excel(output_file, index=False)
            
        else:
            # Chunked processing
            print(f"  Strategy: Chunked read (chunk size: {chunk_size})")
            
            all_chunks = []
            skip_rows = 0
            chunk_num = 0
            
            while True:
                if skip_rows > 0:
                    df_chunk = pd.read_excel(
                        input_file,
                        skiprows=range(1, skip_rows + 1),
                        nrows=chunk_size
                    )
                else:
                    df_chunk = pd.read_excel(input_file, nrows=chunk_size)
                
                if df_chunk.empty:
                    break
                
                chunk_num += 1
                
                # Process chunk
                df_chunk['Total'] = df_chunk['Sales'] * df_chunk['Quantity']
                all_chunks.append(df_chunk)
                
                print(f"  Processed chunk {chunk_num}: {len(df_chunk):,} rows")
                
                skip_rows += chunk_size
                
                if len(df_chunk) < chunk_size:
                    break
            
            # Combine and save
            df_final = pd.concat(all_chunks, ignore_index=True)
            df_final.to_excel(output_file, index=False)
            
        print(f"  ✓ Success: Saved to {Path(output_file).name}")
        return True
        
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return False

# Test production processor
processed_file = output_dir / 'processed_large.xlsx'
success = production_file_processor(large_file, processed_file)
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 80)
print("SESSION 9 COMPLETE!")
print("=" * 80)
print()
print("What you learned:")
print("✓ Measure and categorize file sizes")
print("✓ Read files in chunks")
print("✓ Use openpyxl streaming (read-only mode)")
print("✓ Write efficiently with write-only mode")
print("✓ Optimize data types for memory savings")
print("✓ Read only needed columns")
print("✓ Convert to CSV for faster processing")
print("✓ Process multiple files in batch")
print("✓ Build production-ready file processors")
print("✓ Handle large files without memory issues")
print()
print("Files created:")
for file in sorted(output_dir.glob('*.xlsx')):
    print(f"  - {file.name}")
print()
print("Performance Tips:")
print("  - Use chunking for files > 50MB")
print("  - Use streaming for files > 200MB")
print("  - Convert to CSV when possible")
print("  - Optimize data types")
print("  - Read only needed columns")
print()
print("Next: Session 10 - Automation and Batch Processing")
print("=" * 80)

