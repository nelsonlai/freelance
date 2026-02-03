"""
Session 2: Reading Excel Files - Comprehensive Examples
=======================================================

This script demonstrates advanced techniques for reading Excel files
with various structures and requirements.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
import sys

print("=" * 80)
print("Session 2: Reading Excel Files - Comprehensive Techniques")
print("=" * 80)
print()

# Create sample data directory
sample_dir = Path(__file__).parent / 'sample_files'
sample_dir.mkdir(exist_ok=True)

# ==============================================================================
# 1. CREATE VARIOUS SAMPLE FILES FOR DEMONSTRATION
# ==============================================================================

print("1. CREATING SAMPLE EXCEL FILES WITH DIFFERENT STRUCTURES")
print("-" * 80)

# Sample 1: Simple dataset
df_simple = pd.DataFrame({
    'Name': ['Alice', 'Bob', 'Charlie', 'Diana', 'Eve'],
    'Age': [25, 30, 35, 28, 32],
    'Department': ['Sales', 'IT', 'Sales', 'HR', 'IT'],
    'Salary': [60000, 75000, 65000, 58000, 72000]
})

simple_file = sample_dir / 'simple_data.xlsx'
df_simple.to_excel(simple_file, index=False)
print(f"✓ Created: {simple_file}")

# Sample 2: Multi-sheet workbook
df_sales = pd.DataFrame({
    'Product': ['A', 'B', 'C', 'D'],
    'Q1': [100, 150, 200, 175],
    'Q2': [110, 160, 210, 180],
    'Q3': [120, 170, 220, 190],
    'Q4': [130, 180, 230, 200]
})

df_expenses = pd.DataFrame({
    'Category': ['Rent', 'Utilities', 'Salaries', 'Marketing'],
    'Amount': [5000, 1500, 25000, 3000]
})

multi_sheet_file = sample_dir / 'multi_sheet.xlsx'
with pd.ExcelWriter(multi_sheet_file) as writer:
    df_sales.to_excel(writer, sheet_name='Sales', index=False)
    df_expenses.to_excel(writer, sheet_name='Expenses', index=False)

print(f"✓ Created: {multi_sheet_file}")

# Sample 3: File with headers in row 3
df_with_header_offset = pd.DataFrame({
    'ID': range(1, 11),
    'Product': [f'Product_{i}' for i in range(1, 11)],
    'Price': np.random.uniform(10, 100, 10).round(2)
})

header_offset_file = sample_dir / 'header_offset.xlsx'
with pd.ExcelWriter(header_offset_file, engine='openpyxl') as writer:
    # Write DataFrame starting from row 3
    df_with_header_offset.to_excel(writer, sheet_name='Data', 
                                   startrow=2, index=False)
    
    # Access the worksheet to add headers
    worksheet = writer.sheets['Data']
    worksheet.cell(1, 1, 'Company Sales Report')
    worksheet.cell(2, 1, 'Generated: 2024')

print(f"✓ Created: {header_offset_file}")
print()

# ==============================================================================
# 2. BASIC READING TECHNIQUES
# ==============================================================================

print("2. BASIC READING TECHNIQUES")
print("-" * 80)

# Read simple file
df = pd.read_excel(simple_file)
print("Simple file read:")
print(df)
print()

# Get basic info
print("DataFrame Info:")
print(f"Shape: {df.shape}")
print(f"Columns: {df.columns.tolist()}")
print(f"Data types:\n{df.dtypes}")
print()

# ==============================================================================
# 3. READING SPECIFIC SHEETS
# ==============================================================================

print("3. READING SPECIFIC SHEETS")
print("-" * 80)

# Read specific sheet by name
df_sales_read = pd.read_excel(multi_sheet_file, sheet_name='Sales')
print("Sales sheet:")
print(df_sales_read)
print()

# Read specific sheet by index
df_expenses_read = pd.read_excel(multi_sheet_file, sheet_name=1)
print("Expenses sheet (by index):")
print(df_expenses_read)
print()

# Read all sheets
all_sheets = pd.read_excel(multi_sheet_file, sheet_name=None)
print(f"All sheets: {list(all_sheets.keys())}")
for sheet_name, df_sheet in all_sheets.items():
    print(f"\n{sheet_name}:")
    print(df_sheet.head())
print()

# ==============================================================================
# 4. READING SPECIFIC ROWS AND COLUMNS
# ==============================================================================

print("4. READING SPECIFIC ROWS AND COLUMNS")
print("-" * 80)

# Skip rows
df_skip = pd.read_excel(simple_file, skiprows=2)
print("Skip first 2 data rows:")
print(df_skip)
print()

# Read specific columns
df_cols = pd.read_excel(simple_file, usecols=['Name', 'Salary'])
print("Read only Name and Salary columns:")
print(df_cols)
print()

# Read specific columns by index
df_cols_idx = pd.read_excel(simple_file, usecols=[0, 3])
print("Read columns 0 and 3:")
print(df_cols_idx)
print()

# Limit number of rows
df_nrows = pd.read_excel(simple_file, nrows=3)
print("Read only first 3 rows:")
print(df_nrows)
print()

# ==============================================================================
# 5. HANDLING HEADERS
# ==============================================================================

print("5. HANDLING HEADERS")
print("-" * 80)

# Read with header in different row
df_offset = pd.read_excel(header_offset_file, sheet_name='Data', header=2)
print("Read with header in row 3:")
print(df_offset)
print()

# Read without headers
df_no_header = pd.read_excel(simple_file, header=None)
print("Read without using headers:")
print(df_no_header.head())
print()

# Read without headers and assign custom names
df_custom_names = pd.read_excel(simple_file, header=None, 
                                names=['Col1', 'Col2', 'Col3', 'Col4'],
                                skiprows=1)
print("Custom column names:")
print(df_custom_names.head())
print()

# ==============================================================================
# 6. DATA TYPE HANDLING
# ==============================================================================

print("6. DATA TYPE HANDLING")
print("-" * 80)

# Create file with mixed types
df_mixed = pd.DataFrame({
    'ID': ['001', '002', '003', '004'],
    'Name': ['Alice', 'Bob', 'Charlie', 'Diana'],
    'Date': pd.date_range('2024-01-01', periods=4),
    'Score': [95.5, 87.3, 92.1, 88.9],
    'Active': [True, True, False, True]
})

mixed_file = sample_dir / 'mixed_types.xlsx'
df_mixed.to_excel(mixed_file, index=False)

# Read with automatic type inference
df_auto = pd.read_excel(mixed_file)
print("Automatic type inference:")
print(df_auto.dtypes)
print()

# Read with explicit types
df_explicit = pd.read_excel(mixed_file, dtype={
    'ID': str,
    'Name': str,
    'Score': float,
    'Active': bool
})
print("Explicit type specification:")
print(df_explicit.dtypes)
print()

# ==============================================================================
# 7. READING WITH OPENPYXL
# ==============================================================================

print("7. READING WITH OPENPYXL")
print("-" * 80)

# Load workbook
wb = load_workbook(simple_file)
print(f"Workbook sheets: {wb.sheetnames}")
print()

# Access sheet
ws = wb.active
print(f"Active sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print()

# Read specific cells
print("Reading specific cells:")
print(f"A1 (header): {ws['A1'].value}")
print(f"A2 (first data): {ws['A2'].value}")
print(f"D2 (salary): {ws['D2'].value}")
print()

# Iterate through rows
print("First 3 rows:")
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
    print(f"Row {idx}: {row}")
print()

# Read specific range
print("Reading range A1:C3:")
for row in ws['A1:C3']:
    values = [cell.value for cell in row]
    print(values)
print()

# ==============================================================================
# 8. COMBINING MULTIPLE SHEETS
# ==============================================================================

print("8. COMBINING MULTIPLE SHEETS")
print("-" * 80)

# Create multiple sheet files
monthly_data = {}
for month in ['January', 'February', 'March']:
    monthly_data[month] = pd.DataFrame({
        'Product': ['A', 'B', 'C'],
        'Sales': np.random.randint(100, 500, 3),
        'Month': month
    })

monthly_file = sample_dir / 'monthly_sales.xlsx'
with pd.ExcelWriter(monthly_file) as writer:
    for month, df_month in monthly_data.items():
        df_month.to_excel(writer, sheet_name=month, index=False)

print(f"✓ Created monthly sales file: {monthly_file}")
print()

# Read and combine all sheets
all_months = pd.read_excel(monthly_file, sheet_name=None)
combined_df = pd.concat(all_months.values(), ignore_index=True)
print("Combined data from all months:")
print(combined_df)
print()

# ==============================================================================
# 9. ERROR HANDLING
# ==============================================================================

print("9. ERROR HANDLING BEST PRACTICES")
print("-" * 80)

def safe_read_excel(file_path, **kwargs):
    """Safely read Excel file with error handling"""
    try:
        file_path = Path(file_path)
        
        if not file_path.exists():
            print(f"✗ Error: File not found: {file_path}")
            return None
        
        if not str(file_path).endswith(('.xlsx', '.xls', '.xlsm')):
            print(f"✗ Error: Invalid file format")
            return None
        
        df = pd.read_excel(file_path, **kwargs)
        
        if df.empty:
            print(f"⚠ Warning: File is empty")
            return None
        
        print(f"✓ Successfully read: {file_path.name}")
        return df
        
    except PermissionError:
        print(f"✗ Error: Permission denied. File may be open in Excel.")
        return None
    
    except ValueError as e:
        print(f"✗ ValueError: {e}")
        return None
    
    except Exception as e:
        print(f"✗ Unexpected error: {type(e).__name__}: {e}")
        return None

# Test error handling
print("Testing error handling:")
df = safe_read_excel(simple_file)
if df is not None:
    print(f"DataFrame shape: {df.shape}")
print()

df = safe_read_excel('non_existent.xlsx')
print()

df = safe_read_excel(simple_file, sheet_name='NonExistentSheet')
print()

# ==============================================================================
# 10. PERFORMANCE COMPARISON
# ==============================================================================

print("10. PERFORMANCE COMPARISON")
print("-" * 80)

# Create larger dataset
large_data = pd.DataFrame({
    'ID': range(1, 1001),
    'Name': [f'Person_{i}' for i in range(1, 1001)],
    'Value': np.random.randn(1000)
})

large_file = sample_dir / 'large_dataset.xlsx'
large_data.to_excel(large_file, index=False)

import time

# Test pandas
start = time.time()
df_pandas = pd.read_excel(large_file)
pandas_time = time.time() - start

# Test openpyxl read_only
start = time.time()
wb = load_workbook(large_file, read_only=True)
ws = wb.active
data_list = [[cell.value for cell in row] for row in ws.iter_rows()]
df_openpyxl = pd.DataFrame(data_list[1:], columns=data_list[0])
wb.close()
openpyxl_time = time.time() - start

print(f"Pandas read time: {pandas_time:.4f} seconds")
print(f"Openpyxl read time: {openpyxl_time:.4f} seconds")
print()

if pandas_time < openpyxl_time:
    print(f"✓ Pandas is {openpyxl_time/pandas_time:.2f}x faster")
else:
    print(f"✓ Openpyxl is {pandas_time/openpyxl_time:.2f}x faster")
print()

# ==============================================================================
# SUMMARY
# ==============================================================================

print("=" * 80)
print("SESSION 2 COMPLETE!")
print("=" * 80)
print()
print("What you learned:")
print("✓ Read single and multiple sheets")
print("✓ Read specific rows, columns, and ranges")
print("✓ Handle different data types")
print("✓ Work with headers and indices")
print("✓ Use both pandas and openpyxl effectively")
print("✓ Implement error handling")
print("✓ Compare performance of different methods")
print()
print("Files created:")
for file in sample_dir.glob('*.xlsx'):
    print(f"  - {file.name}")
print()
print("Next: Session 3 - Writing Excel Files")
print("=" * 80)

