"""
Session 12: Real-World Projects and Best Practices
==================================================

This script demonstrates three complete real-world Excel automation projects:
1. Financial Report Generator
2. Sales Data Dashboard
3. ETL Pipeline for Excel Data

Each project showcases production-ready code with best practices.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from datetime import datetime, timedelta
import logging

print("=" * 80)
print("Session 12: Real-World Projects")
print("=" * 80)
print()

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Create output directory
output_dir = Path(__file__).parent / 'generated_reports'
output_dir.mkdir(exist_ok=True)

# ==============================================================================
# PROJECT 1: FINANCIAL REPORT GENERATOR
# ==============================================================================

print("PROJECT 1: FINANCIAL REPORT GENERATOR")
print("-" * 80)

class FinancialReportGenerator:
    """
    Professional financial report generator with formatting and charts.
    """
    
    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def generate_sample_data(self):
        """Generate sample financial data"""
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
        
        data = {
            'Month': months,
            'Revenue': [120000, 135000, 128000, 142000, 155000, 148000],
            'COGS': [48000, 54000, 51200, 56800, 62000, 59200],
            'Operating_Expenses': [35000, 36000, 34500, 37000, 38500, 37500],
            'Marketing': [15000, 18000, 16000, 19000, 20000, 19500],
        }
        
        df = pd.DataFrame(data)
        
        # Calculate derived metrics
        df['Gross_Profit'] = df['Revenue'] - df['COGS']
        df['Total_Expenses'] = df['Operating_Expenses'] + df['Marketing']
        df['Net_Income'] = df['Gross_Profit'] - df['Total_Expenses']
        df['Profit_Margin'] = (df['Net_Income'] / df['Revenue'] * 100).round(2)
        
        return df
    
    def create_report(self, data, filename):
        """Create formatted financial report"""
        try:
            filepath = self.output_dir / filename
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Report"
            
            # Add title
            ws['A1'] = 'Financial Performance Report'
            ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:I1')
            
            # Add timestamp
            ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            ws['A2'].font = Font(italic=True)
            ws.merge_cells('A2:I2')
            
            # Write headers
            headers = list(data.columns)
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col_num, value=header.replace('_', ' '))
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Write data
            for row_idx, row in enumerate(data.itertuples(index=False), start=5):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format currency columns
                    if col_idx in [2, 3, 4, 5, 6, 7, 8]:
                        if col_idx == 9:  # Profit margin (percentage)
                            cell.number_format = '0.00"%"'
                        else:
                            cell.number_format = '$#,##0'
                    
                    # Add borders
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Add summary section
            summary_row = len(data) + 6
            ws.cell(summary_row, 1, 'TOTALS:').font = Font(bold=True)
            
            for col in [2, 3, 4, 5, 6, 7, 8]:
                col_letter = chr(64 + col)
                formula = f'=SUM({col_letter}5:{col_letter}{len(data)+4})'
                cell = ws.cell(summary_row, col, formula)
                cell.number_format = '$#,##0'
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            # Add chart
            chart = BarChart()
            chart.title = "Revenue vs Net Income"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Month"
            
            # Revenue data
            revenue_data = Reference(ws, min_col=2, min_row=4, max_row=len(data)+4)
            # Net Income data
            net_income_data = Reference(ws, min_col=8, min_row=4, max_row=len(data)+4)
            # Categories (months)
            cats = Reference(ws, min_col=1, min_row=5, max_row=len(data)+4)
            
            chart.add_data(revenue_data, titles_from_data=True)
            chart.add_data(net_income_data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, f'A{summary_row + 3}')
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            wb.save(filepath)
            self.logger.info(f"✓ Financial report created: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating report: {str(e)}")
            return False

# Run Project 1
print("Generating financial report...")
generator = FinancialReportGenerator(output_dir)
financial_data = generator.generate_sample_data()
generator.create_report(financial_data, 'financial_report.xlsx')
print()

# ==============================================================================
# PROJECT 2: SALES DATA DASHBOARD
# ==============================================================================

print("PROJECT 2: SALES DATA DASHBOARD")
print("-" * 80)

class SalesDashboard:
    """
    Interactive sales dashboard with multiple charts and KPIs.
    """
    
    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def generate_sales_data(self):
        """Generate sample sales data"""
        np.random.seed(42)
        
        dates = pd.date_range(start='2024-01-01', end='2024-06-30', freq='D')
        regions = ['North', 'South', 'East', 'West']
        products = ['Product A', 'Product B', 'Product C', 'Product D']
        
        data = []
        for date in dates:
            for region in regions:
                for product in products:
                    quantity = np.random.randint(10, 100)
                    price = np.random.uniform(50, 500)
                    data.append({
                        'Date': date,
                        'Region': region,
                        'Product': product,
                        'Quantity': quantity,
                        'Unit_Price': round(price, 2),
                        'Total_Sales': round(quantity * price, 2)
                    })
        
        return pd.DataFrame(data)
    
    def create_dashboard(self, data, filename):
        """Create sales dashboard with KPIs and charts"""
        try:
            filepath = self.output_dir / filename
            
            # Create Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Summary sheet
                summary = self.create_summary(data)
                summary.to_excel(writer, sheet_name='Summary', index=False)
                
                # By Region
                by_region = data.groupby('Region')['Total_Sales'].sum().reset_index()
                by_region.to_excel(writer, sheet_name='By Region', index=False)
                
                # By Product
                by_product = data.groupby('Product')['Total_Sales'].sum().reset_index()
                by_product.to_excel(writer, sheet_name='By Product', index=False)
                
                # Daily Sales
                daily = data.groupby('Date')['Total_Sales'].sum().reset_index()
                daily.to_excel(writer, sheet_name='Daily Sales', index=False)
                
                # Raw Data
                data.to_excel(writer, sheet_name='Raw Data', index=False)
            
            # Add charts
            self.add_charts(filepath)
            
            self.logger.info(f"✓ Sales dashboard created: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating dashboard: {str(e)}")
            return False
    
    def create_summary(self, data):
        """Create KPI summary"""
        summary_data = {
            'Metric': [
                'Total Sales',
                'Average Daily Sales',
                'Total Transactions',
                'Average Transaction Value',
                'Top Region',
                'Top Product'
            ],
            'Value': [
                f"${data['Total_Sales'].sum():,.2f}",
                f"${data.groupby('Date')['Total_Sales'].sum().mean():,.2f}",
                len(data),
                f"${data['Total_Sales'].mean():,.2f}",
                data.groupby('Region')['Total_Sales'].sum().idxmax(),
                data.groupby('Product')['Total_Sales'].sum().idxmax()
            ]
        }
        return pd.DataFrame(summary_data)
    
    def add_charts(self, filepath):
        """Add charts to dashboard"""
        wb = load_workbook(filepath)
        
        # Add chart to By Region sheet
        ws_region = wb['By Region']
        chart = PieChart()
        chart.title = "Sales by Region"
        data = Reference(ws_region, min_col=2, min_row=1, max_row=5)
        cats = Reference(ws_region, min_col=1, min_row=2, max_row=5)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_region.add_chart(chart, "D2")
        
        # Add chart to By Product sheet
        ws_product = wb['By Product']
        chart2 = BarChart()
        chart2.title = "Sales by Product"
        chart2.y_axis.title = "Total Sales"
        data2 = Reference(ws_product, min_col=2, min_row=1, max_row=5)
        cats2 = Reference(ws_product, min_col=1, min_row=2, max_row=5)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        ws_product.add_chart(chart2, "D2")
        
        wb.save(filepath)

# Run Project 2
print("Generating sales dashboard...")
dashboard = SalesDashboard(output_dir)
sales_data = dashboard.generate_sales_data()
dashboard.create_dashboard(sales_data, 'sales_dashboard.xlsx')
print()

# ==============================================================================
# PROJECT 3: ETL PIPELINE FOR EXCEL DATA
# ==============================================================================

print("PROJECT 3: ETL PIPELINE")
print("-" * 80)

class ExcelETLPipeline:
    """
    Extract-Transform-Load pipeline for Excel data processing.
    """
    
    def __init__(self, input_dir, output_dir):
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logging.getLogger(__name__)
    
    def extract(self, pattern='*.xlsx'):
        """Extract data from multiple Excel files"""
        files = list(self.input_dir.glob(pattern))
        
        if not files:
            self.logger.warning(f"No files found matching pattern: {pattern}")
            return pd.DataFrame()
        
        all_data = []
        for file in files:
            try:
                df = pd.read_excel(file)
                df['source_file'] = file.name
                all_data.append(df)
                self.logger.info(f"✓ Extracted: {file.name}")
            except Exception as e:
                self.logger.error(f"✗ Failed to extract {file.name}: {str(e)}")
        
        if all_data:
            return pd.concat(all_data, ignore_index=True)
        return pd.DataFrame()
    
    def transform(self, data):
        """Transform and clean data"""
        if data.empty:
            return data
        
        # Remove duplicates
        original_rows = len(data)
        data = data.drop_duplicates()
        self.logger.info(f"Removed {original_rows - len(data)} duplicate rows")
        
        # Handle missing values
        data = data.ffill()
        
        # Add derived columns
        if 'Date' in data.columns:
            data['Date'] = pd.to_datetime(data['Date'])
            data['Year'] = data['Date'].dt.year
            data['Month'] = data['Date'].dt.month
            data['Quarter'] = data['Date'].dt.quarter
        
        # Standardize column names
        data.columns = [col.lower().replace(' ', '_') for col in data.columns]
        
        self.logger.info(f"✓ Transformation complete: {len(data)} rows")
        return data
    
    def load(self, data, filename):
        """Load processed data to Excel"""
        if data.empty:
            self.logger.warning("No data to load")
            return False
        
        try:
            filepath = self.output_dir / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Main data
                data.to_excel(writer, sheet_name='Processed Data', index=False)
                
                # Summary statistics
                if data.select_dtypes(include=[np.number]).columns.any():
                    summary = data.describe()
                    summary.to_excel(writer, sheet_name='Statistics')
                
                # Create pivot if date columns exist
                if 'date' in data.columns or 'year' in data.columns:
                    try:
                        numeric_cols = data.select_dtypes(include=[np.number]).columns
                        if len(numeric_cols) > 0 and 'month' in data.columns:
                            pivot = data.pivot_table(
                                values=numeric_cols[0],
                                index='month',
                                aggfunc='sum'
                            )
                            pivot.to_excel(writer, sheet_name='Monthly Summary')
                    except Exception as e:
                        self.logger.debug(f"Could not create pivot: {str(e)}")
            
            self.logger.info(f"✓ Data loaded to: {filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading data: {str(e)}")
            return False
    
    def run_pipeline(self, output_filename='etl_output.xlsx'):
        """Run the complete ETL pipeline"""
        self.logger.info("Starting ETL pipeline...")
        
        # Extract
        data = self.extract()
        if data.empty:
            self.logger.error("No data extracted. Pipeline stopped.")
            return False
        
        # Transform
        data = self.transform(data)
        
        # Load
        success = self.load(data, output_filename)
        
        if success:
            self.logger.info("✓ ETL pipeline completed successfully!")
        else:
            self.logger.error("✗ ETL pipeline failed")
        
        return success

# Run Project 3 (demo mode - creates sample data first)
print("Creating sample data for ETL pipeline...")
sample_dir = output_dir / 'etl_input'
sample_dir.mkdir(exist_ok=True)

# Create sample input files
for i in range(3):
    sample_df = pd.DataFrame({
        'Date': pd.date_range(start='2024-01-01', periods=30),
        'Product': [f'Product {j%3}' for j in range(30)],
        'Sales': np.random.randint(100, 1000, 30),
        'Quantity': np.random.randint(1, 50, 30)
    })
    sample_df.to_excel(sample_dir / f'sample_data_{i+1}.xlsx', index=False)

print("Running ETL pipeline...")
pipeline = ExcelETLPipeline(sample_dir, output_dir)
pipeline.run_pipeline('etl_processed_data.xlsx')
print()

# ==============================================================================
# SUMMARY AND BEST PRACTICES
# ==============================================================================

print("=" * 80)
print("SESSION 12 COMPLETE!")
print("=" * 80)
print()
print("Projects completed:")
print("✓ 1. Financial Report Generator - Formatted reports with charts")
print("✓ 2. Sales Data Dashboard - Multi-sheet analytics dashboard")
print("✓ 3. ETL Pipeline - Complete data processing workflow")
print()
print("Key best practices demonstrated:")
print("✓ Object-oriented design with classes")
print("✓ Logging and error handling")
print("✓ Input validation and data cleaning")
print("✓ Professional formatting and styling")
print("✓ Automated chart generation")
print("✓ Modular, reusable code")
print("✓ Documentation and comments")
print()
print(f"All reports saved to: {output_dir}")
print()
print("Congratulations! You've completed the Python Excel Operations course!")
print("=" * 80)

