#!/usr/bin/env python3
"""
LeetCode Problem Downloader
Downloads all problems from LeetCode and saves them to Excel and PDF files.
"""

import requests
import json
import os
import time
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Image as RLImage
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from html2text import HTML2Text
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import re
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import load_workbook
from solution_generator import SolutionGenerator


class LeetCodeDownloader:
    """Downloads LeetCode problems and generates Excel and PDF files."""
    
    def __init__(self, output_dir: str = "leetcode_problems", session_cookie: str = None):
        self.output_dir = output_dir
        self.pdf_dir = os.path.join(output_dir, "pdfs")
        self.images_dir = os.path.join(output_dir, "images")
        self.codes_dir = os.path.join(output_dir, "codes")
        self.graphql_url = "https://leetcode.com/graphql/"
        self.headers = {
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        
        # Add authentication if provided
        if session_cookie:
            self.headers["Cookie"] = f"LEETCODE_SESSION={session_cookie}"
        
        # Create output directories
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.pdf_dir, exist_ok=True)
        os.makedirs(self.images_dir, exist_ok=True)
        os.makedirs(self.codes_dir, exist_ok=True)
        
        # Initialize solution generator
        self.solution_generator = SolutionGenerator()
        
    def fetch_all_problems(self) -> List[Dict]:
        """Fetch all problems from LeetCode using GraphQL API."""
        print("Fetching all problems from LeetCode...")
        
        query = """
        query problemsetQuestionList($categorySlug: String, $limit: Int, $skip: Int, $filters: QuestionListFilterInput) {
          problemsetQuestionList: questionList(
            categorySlug: $categorySlug
            limit: $limit
            skip: $skip
            filters: $filters
          ) {
            total: totalNum
            questions: data {
              acRate
              difficulty
              freqBar
              frontendQuestionId: questionFrontendId
              isFavor
              paidOnly: isPaidOnly
              status
              title
              titleSlug
              topicTags {
                name
                id
                slug
              }
              hasSolution
              hasVideoSolution
            }
          }
        }
        """
        
        variables = {
            "categorySlug": "",
            "skip": 0,
            "limit": 50,
            "filters": {}
        }
        
        all_problems = []
        skip = 0
        limit = 50
        
        while True:
            variables["skip"] = skip
            payload = {
                "query": query,
                "variables": variables
            }
            
            try:
                response = requests.post(
                    self.graphql_url,
                    headers=self.headers,
                    json=payload,
                    timeout=30
                )
                response.raise_for_status()
                data = response.json()
                
                if "errors" in data:
                    print(f"GraphQL Error: {data['errors']}")
                    break
                
                questions = data.get("data", {}).get("problemsetQuestionList", {}).get("questions", [])
                if not questions:
                    break
                
                all_problems.extend(questions)
                total = data.get("data", {}).get("problemsetQuestionList", {}).get("total", 0)
                
                print(f"Fetched {len(all_problems)}/{total} problems...")
                
                if len(questions) < limit:
                    break
                
                skip += limit
                time.sleep(0.5)  # Be respectful to the API
                
            except requests.exceptions.RequestException as e:
                print(f"Error fetching problems: {e}")
                break
        
        print(f"Total problems fetched: {len(all_problems)}")
        return all_problems
    
    def fetch_problem_detail(self, title_slug: str, max_retries: int = 3) -> Optional[Dict]:
        """Fetch detailed information for a specific problem with retry logic."""
        query = """
        query questionContent($titleSlug: String!) {
          question(titleSlug: $titleSlug) {
            content
            mysqlSchemas
            dataSchemas
            questionId
            questionFrontendId
            title
            titleSlug
            difficulty
            likes
            dislikes
            isLiked
            similarQuestions
            contributors {
              username
              profileUrl
              avatarUrl
              __typename
            }
            topicTags {
              name
              slug
              translatedName
              __typename
            }
            companyTagStats
            codeSnippets {
              lang
              langSlug
              code
              __typename
            }
            stats
            hints
            solution {
              id
              canSeeDetail
              paidOnly
              __typename
            }
            status
            sampleTestCase
            metaData
            judgerAvailable
            judgeType
            mysqlSchemas
            enableRunCode
            enableTestMode
            enableDebugger
            envInfo
            libraryUrl
            adminUrl
            challengeQuestion {
              id
              date
              incompleteChallengeCount
              streakCount
              type
              __typename
            }
            __typename
          }
        }
        """
        
        variables = {"titleSlug": title_slug}
        payload = {
            "query": query,
            "variables": variables
        }
        
        for attempt in range(max_retries):
            try:
                # Increase timeout for retries
                timeout = 60 if attempt > 0 else 30
                response = requests.post(
                    self.graphql_url,
                    headers=self.headers,
                    json=payload,
                    timeout=timeout
                )
                response.raise_for_status()
                data = response.json()
                
                if "errors" in data:
                    return None
                
                return data.get("data", {}).get("question")
            except requests.exceptions.Timeout:
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2
                    print(f"Timeout fetching {title_slug}, retrying in {wait_time}s... (attempt {attempt + 1}/{max_retries})")
                    time.sleep(wait_time)
                else:
                    print(f"Error fetching problem detail for {title_slug}: Read timeout after {max_retries} attempts")
                    return None
            except requests.exceptions.RequestException as e:
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2
                    print(f"Error fetching {title_slug}, retrying in {wait_time}s... (attempt {attempt + 1}/{max_retries})")
                    time.sleep(wait_time)
                else:
                    print(f"Error fetching problem detail for {title_slug}: {e}")
                    return None
        
        return None
    
    def extract_companies(self, company_tag_stats: Optional[str]) -> str:
        """Extract company names from companyTagStats JSON string."""
        if not company_tag_stats:
            return ""
        
        companies = []
        
        try:
            # companyTagStats is a JSON string, parse it
            if isinstance(company_tag_stats, str):
                stats = json.loads(company_tag_stats)
            else:
                stats = company_tag_stats
            
            if not stats:
                return ""
            
            # Handle different possible structures
            if isinstance(stats, list):
                for item in stats:
                    if isinstance(item, dict):
                        # Try different possible field names
                        company_name = None
                        
                        # Check for nested structures
                        if "taggedByAdmin" in item and isinstance(item["taggedByAdmin"], dict):
                            company_name = item["taggedByAdmin"].get("name")
                        elif "taggedBy" in item and isinstance(item["taggedBy"], dict):
                            company_name = item["taggedBy"].get("name")
                        elif "name" in item:
                            company_name = item.get("name")
                        elif "company" in item:
                            company_name = item.get("company")
                            if isinstance(company_name, dict):
                                company_name = company_name.get("name")
                        
                        if company_name:
                            companies.append(company_name)
                    elif isinstance(item, str):
                        companies.append(item)
            elif isinstance(stats, dict):
                # If it's a dict, try to extract company names
                if "name" in stats:
                    companies.append(stats["name"])
                # Check for nested company info
                for key, value in stats.items():
                    if isinstance(value, dict) and "name" in value:
                        companies.append(value["name"])
                    elif isinstance(value, list):
                        for sub_item in value:
                            if isinstance(sub_item, dict) and "name" in sub_item:
                                companies.append(sub_item["name"])
            
            return ", ".join(sorted(set(companies))) if companies else ""
            
        except (json.JSONDecodeError, TypeError, AttributeError, KeyError) as e:
            # If parsing fails, try to extract company names from the string directly
            try:
                if isinstance(company_tag_stats, str):
                    # Look for "name" fields in the JSON string
                    matches = re.findall(r'"name"\s*:\s*"([^"]+)"', company_tag_stats)
                    if matches:
                        return ", ".join(sorted(set(matches)))
            except Exception:
                pass
            return ""
    
    def is_valid_image_url(self, url: str) -> bool:
        """Check if URL is a valid image URL."""
        if not url or not isinstance(url, str):
            return False
        
        # Check if it looks like a URL (starts with http, https, //, or /)
        if not (url.startswith('http://') or url.startswith('https://') or 
                url.startswith('//') or url.startswith('/')):
            return False
        
        # Check if it's not just plain text (no spaces, no obvious text patterns)
        if ' ' in url or '\n' in url or url.lower().startswith('image'):
            return False
        
        # Parse URL to validate
        try:
            parsed = urlparse(url if url.startswith('http') else 'https://example.com' + url)
            # Check if path looks valid
            if parsed.path and not parsed.path.startswith('/'):
                return False
        except Exception:
            return False
        
        return True
    
    def is_supported_image_format(self, filepath: str) -> bool:
        """Check if image format is supported by PIL."""
        supported_formats = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff', '.tif'}
        ext = os.path.splitext(filepath.lower())[1]
        return ext in supported_formats
    
    def download_image(self, image_url: str, problem_id: str, image_index: int) -> Optional[str]:
        """Download an image from URL and save it locally."""
        try:
            # Validate URL first
            if not self.is_valid_image_url(image_url):
                return None
            
            # Convert relative URLs to absolute
            if image_url.startswith('//'):
                image_url = 'https:' + image_url
            elif image_url.startswith('/'):
                image_url = 'https://leetcode.com' + image_url
            
            # Get file extension
            parsed_url = urlparse(image_url)
            ext = os.path.splitext(parsed_url.path)[1].lower()
            
            # Skip unsupported formats (SVG, drawio, etc.)
            unsupported_formats = {'.svg', '.drawio', '.xml'}
            if ext in unsupported_formats:
                return None
            
            # Default to png if no extension
            if not ext:
                ext = '.png'
            
            # Create filename
            filename = f"{problem_id}_img_{image_index}{ext}"
            filepath = os.path.join(self.images_dir, filename)
            
            # Download image
            img_headers = self.headers.copy()
            img_headers.pop("Content-Type", None)  # Remove Content-Type for image requests
            
            response = requests.get(image_url, headers=img_headers, timeout=30, stream=True)
            response.raise_for_status()
            
            # Check content type
            content_type = response.headers.get('content-type', '').lower()
            if 'image' not in content_type and content_type:
                # Not an image, skip it
                return None
            
            # Save image
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            # Verify it's a valid image file that PIL can open
            if not self.is_supported_image_format(filepath):
                # Try to delete unsupported file
                try:
                    os.remove(filepath)
                except:
                    pass
                return None
            
            # Try to open with PIL to verify it's valid
            try:
                PILImage.open(filepath).verify()
            except Exception:
                # Invalid or corrupted image, delete it
                try:
                    os.remove(filepath)
                except:
                    pass
                return None
            
            return filepath
        except Exception as e:
            # Silently skip invalid images (don't print error for every invalid URL)
            return None
    
    def html_to_reportlab(self, html_content: str) -> str:
        """Convert HTML to ReportLab-compatible format while preserving structure."""
        if not html_content:
            return ""
        
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Process different HTML elements
            result_parts = []
            
            # Process all elements in order
            for element in soup.find_all(['p', 'div', 'pre', 'code', 'strong', 'b', 'em', 'i', 
                                         'ul', 'ol', 'li', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
                                         'br', 'img', 'span']):
                if element.name == 'img':
                    # Images are handled separately
                    continue
                elif element.name in ['strong', 'b']:
                    text = element.get_text()
                    if text:
                        result_parts.append(f"<b>{text}</b>")
                elif element.name in ['em', 'i']:
                    text = element.get_text()
                    if text:
                        result_parts.append(f"<i>{text}</i>")
                elif element.name == 'code':
                    text = element.get_text()
                    if text:
                        result_parts.append(f"<font name='Courier'>{text}</font>")
                elif element.name == 'pre':
                    text = element.get_text()
                    if text:
                        result_parts.append(f"<font name='Courier'>{text}</font>")
                elif element.name in ['ul', 'ol']:
                    # Lists are handled by processing li elements
                    continue
                elif element.name == 'li':
                    text = element.get_text(strip=True)
                    if text:
                        result_parts.append(f"• {text}")
                elif element.name == 'br':
                    result_parts.append("<br/>")
                elif element.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                    text = element.get_text(strip=True)
                    if text:
                        result_parts.append(f"<b>{text}</b>")
                else:
                    text = element.get_text(strip=True)
                    if text:
                        result_parts.append(text)
            
            # If no structured content, get all text
            if not result_parts:
                text = soup.get_text(separator=' ', strip=True)
                if text:
                    result_parts.append(text)
            
            return ' '.join(result_parts)
            
        except Exception as e:
            print(f"Error converting HTML to ReportLab format: {e}")
            # Fallback
            text = re.sub(r'<[^>]+>', '', html_content)
            return text.strip()
    
    def parse_html_with_images(self, html_content: str, problem_id: str) -> List[Tuple[str, Optional[str]]]:
        """
        Parse HTML content and extract structured content with images in order.
        Returns a list of tuples: (content_type, content)
        content_type can be 'text', 'image', 'heading', 'code_block'
        Preserves structure like Examples, Constraints, etc.
        """
        if not html_content:
            return []
        
        result = []
        image_counter = 0
        
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Find all images and create placeholders
            img_placeholders = {}
            for img_tag in soup.find_all('img'):
                img_src = img_tag.get('src', '')
                if img_src:
                    placeholder = f"__IMAGE_PLACEHOLDER_{image_counter}__"
                    img_placeholders[placeholder] = img_src
                    img_tag.replace_with(placeholder)
                    image_counter += 1
            
            # Get the main content - try to find the body or use the whole soup
            main_content = soup.find('body') or soup
            
            # Process elements in order, preserving structure
            def convert_to_reportlab_html(elem):
                """Convert HTML element to ReportLab-compatible HTML string."""
                if elem.name is None:
                    return str(elem).strip()
                
                if elem.name in ['strong', 'b']:
                    return f"<b>{elem.get_text()}</b>"
                elif elem.name in ['em', 'i']:
                    return f"<i>{elem.get_text()}</i>"
                elif elem.name == 'code':
                    code_text = elem.get_text()
                    return f"<font name='Courier'>{code_text}</font>"
                else:
                    return elem.get_text()
            
            # Process all elements maintaining order
            for elem in main_content.descendants:
                if not hasattr(elem, 'name'):
                    continue
                
                if elem.name == 'img':
                    # Handle image placeholder
                    if '__IMAGE_PLACEHOLDER_' in str(elem):
                        continue  # Already handled
                
                elif elem.name == 'p':
                    # Process paragraph
                    text_parts = []
                    if hasattr(elem, 'children'):
                        for child in elem.children:
                            if hasattr(child, 'name'):
                                if child.name == 'img':
                                    src = child.get('src', '')
                                    if src in img_placeholders.values():
                                        placeholder = [k for k, v in img_placeholders.items() if v == src][0]
                                        img_index = int(re.search(r'\d+', placeholder).group())
                                        img_path = self.download_image(src, problem_id, img_index)
                                        if img_path:
                                            result.append(('image', img_path))
                                elif child.name in ['strong', 'b', 'em', 'i', 'code']:
                                    text_parts.append(convert_to_reportlab_html(child))
                                else:
                                    text_parts.append(child.get_text())
                            else:
                                text = str(child).strip()
                                if text:
                                    text_parts.append(text)
                    
                    para_text = ' '.join(text_parts).strip()
                    if para_text:
                        result.append(('text', para_text))
                
                elif elem.name == 'pre':
                    # Code block
                    code_text = elem.get_text()
                    if code_text.strip():
                        result.append(('code_block', code_text))
                
                elif elem.name in ['ul', 'ol']:
                    # Lists - process items
                    for idx, li in enumerate(elem.find_all('li', recursive=False), 1):
                        li_text = li.get_text(strip=True)
                        if li_text:
                            prefix = f"{idx}. " if elem.name == 'ol' else "• "
                            result.append(('text', f"{prefix}{li_text}"))
                
                elif elem.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                    # Headings
                    heading_text = elem.get_text(strip=True)
                    if heading_text:
                        result.append(('heading', heading_text))
            
            # Also process direct text nodes and handle images in text
            text_content = main_content.get_text(separator='\n', strip=False)
            
            # Split by image placeholders and process
            if img_placeholders:
                parts = re.split(r'(__IMAGE_PLACEHOLDER_\d+__)', text_content)
                temp_result = []
                for part in parts:
                    if part.startswith('__IMAGE_PLACEHOLDER_') and part.endswith('__'):
                        img_src = img_placeholders.get(part)
                        if img_src:
                            img_index = int(re.search(r'\d+', part).group())
                            img_path = self.download_image(img_src, problem_id, img_index)
                            if img_path:
                                temp_result.append(('image', img_path))
                    elif part.strip():
                        # Clean up the text
                        cleaned = re.sub(r'\n{3,}', '\n\n', part.strip())
                        if cleaned:
                            temp_result.append(('text', cleaned))
                
                # Merge with structured result, prioritizing structured
                if temp_result:
                    # Use structured result if we have it, otherwise use text-based
                    if result:
                        # Merge images from temp_result into result at appropriate positions
                        final_result = []
                        text_combined = ' '.join([str(d) for t, d in temp_result if t == 'text'])
                        img_list = [(t, d) for t, d in temp_result if t == 'image']
                        
                        # Add structured content
                        for content_type, content_data in result:
                            final_result.append((content_type, content_data))
                            # Insert images that might belong here
                            if content_type == 'text' and img_list:
                                # Try to match images with text (simplified)
                                pass
                        
                        # Add remaining images
                        for img_type, img_data in img_list:
                            if (img_type, img_data) not in [(t, d) for t, d in final_result]:
                                final_result.append((img_type, img_data))
                        
                        result = final_result
                    else:
                        result = temp_result
            
            # If still no result, fall back to simple extraction
            if not result:
                # Use html2text for better formatting
                try:
                    h = HTML2Text()
                    h.ignore_links = True
                    h.ignore_images = True
                    h.body_width = 0  # Don't wrap
                    text = h.handle(html_content)
                    # Clean up
                    text = re.sub(r'\n{3,}', '\n\n', text)
                    paragraphs = text.split('\n\n')
                    for para in paragraphs:
                        if para.strip():
                            result.append(('text', para.strip()))
                except:
                    # Final fallback
                    text = soup.get_text(separator='\n\n', strip=True)
                    if text:
                        result.append(('text', text))
            
            return result
                    
        except Exception as e:
            print(f"Error parsing HTML: {e}")
            # Fallback to simple text extraction
            try:
                h = HTML2Text()
                h.ignore_links = True
                h.ignore_images = True
                text = h.handle(html_content)
                text = re.sub(r'\n{3,}', '\n\n', text)
                paragraphs = text.split('\n\n')
                result = []
                for para in paragraphs:
                    if para.strip():
                        result.append(('text', para.strip()))
                return result
            except:
                text = re.sub(r'<[^>]+>', '', html_content)
                if text.strip():
                    return [('text', text.strip())]
                return []
    
    def html_to_text(self, html_content: str) -> str:
        """Convert HTML content to plain text with better formatting."""
        if not html_content:
            return ""
        try:
            h = HTML2Text()
            h.ignore_links = False
            h.ignore_images = True
            h.body_width = 0  # Don't wrap lines
            h.unicode_snob = True
            h.skip_internal_links = True
            text = h.handle(html_content)
            # Clean up the text
            text = re.sub(r'\n{3,}', '\n\n', text)
            return text.strip()
        except Exception as e:
            print(f"Error converting HTML to text: {e}")
            # Fallback: basic HTML tag removal
            text = re.sub(r'<[^>]+>', '', html_content)
            return text.strip()
    
    def parse_html_to_structured_content(self, html_content: str, problem_id: str) -> List[Tuple[str, str]]:
        """
        Parse HTML and convert to structured content preserving format.
        Returns list of (content_type, content) where content_type is 'paragraph', 'heading', 'code', 'image', 'list'
        """
        if not html_content:
            return []
        
        result = []
        image_counter = 0
        
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # Handle images first
            img_placeholders = {}
            for img_tag in soup.find_all('img'):
                img_src = img_tag.get('src', '')
                if img_src:
                    placeholder = f"__IMG_{image_counter}__"
                    img_placeholders[placeholder] = img_src
                    img_tag.replace_with(placeholder)
                    image_counter += 1
            
            # Process content element by element
            def process_node(node, in_list=False):
                """Process a node and its children."""
                nonlocal result, img_placeholders
                
                if not hasattr(node, 'name'):
                    return ""
                
                if node.name == 'p':
                    # Paragraph
                    text_parts = []
                    if hasattr(node, 'children'):
                        for child in node.children:
                            if hasattr(child, 'name'):
                                if child.name == 'img' or '__IMG_' in str(child):
                                    continue  # Images handled separately
                                elif child.name in ['strong', 'b']:
                                    text_parts.append(f"<b>{child.get_text()}</b>")
                                elif child.name in ['em', 'i']:
                                    text_parts.append(f"<i>{child.get_text()}</i>")
                                elif child.name == 'code':
                                    text_parts.append(f"<font name='Courier'>{child.get_text()}</font>")
                                else:
                                    child_text = process_node(child, in_list)
                                    if child_text:
                                        text_parts.append(child_text)
                            else:
                                text = str(child).strip()
                                if text:
                                    text_parts.append(text)
                    
                    para_text = ' '.join(text_parts).strip()
                    if para_text:
                        result.append(('paragraph', para_text))
                    return ""
                
                elif node.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                    heading_text = node.get_text(strip=True)
                    if heading_text:
                        result.append(('heading', heading_text))
                    return ""
                
                elif node.name == 'pre':
                    code_text = node.get_text()
                    if code_text.strip():
                        result.append(('code', code_text))
                    return ""
                
                elif node.name == 'code' and not in_list:
                    return f"<font name='Courier'>{node.get_text()}</font>"
                
                elif node.name == 'ul':
                    for li in node.find_all('li', recursive=False):
                        li_text = li.get_text(strip=True)
                        if li_text:
                            result.append(('list_item', f"• {li_text}"))
                    return ""
                
                elif node.name == 'ol':
                    for idx, li in enumerate(node.find_all('li', recursive=False), 1):
                        li_text = li.get_text(strip=True)
                        if li_text:
                            result.append(('list_item', f"{idx}. {li_text}"))
                    return ""
                
                elif node.name == 'li':
                    return node.get_text(strip=True)
                
                elif node.name == 'br':
                    result.append(('paragraph', '<br/>'))
                    return ""
                
                else:
                    # Process children - only if node is a Tag (not NavigableString)
                    if hasattr(node, 'children') and hasattr(node, 'name'):
                        parts = []
                        for child in node.children:
                            if hasattr(child, 'name'):
                                part = process_node(child, in_list or node.name in ['ul', 'ol'])
                                if part:
                                    parts.append(part)
                            else:
                                text = str(child).strip()
                                if text:
                                    parts.append(text)
                        return ' '.join(parts)
                    else:
                        # For NavigableString or other non-Tag nodes, just return the text
                        return node.get_text() if hasattr(node, 'get_text') else str(node).strip()
            
            # Process main content
            main = soup.find('body') or soup
            if hasattr(main, 'children'):
                for child in main.children:
                    if hasattr(child, 'name'):
                        process_node(child)
                    elif hasattr(child, 'strip'):  # NavigableString
                        text = str(child).strip()
                        if text:
                            result.append(('paragraph', text))
            
            # Handle images - insert them at appropriate positions
            # For now, append images at the end (can be improved)
            for placeholder, img_src in img_placeholders.items():
                img_index = int(re.search(r'\d+', placeholder).group())
                img_path = self.download_image(img_src, problem_id, img_index)
                if img_path:
                    result.append(('image', img_path))
            
            # If no structured content, use html2text
            if not result:
                text = self.html_to_text(html_content)
                if text:
                    paragraphs = text.split('\n\n')
                    for para in paragraphs:
                        if para.strip():
                            result.append(('paragraph', para.strip()))
            
            return result
            
        except Exception as e:
            print(f"Error parsing HTML to structured content: {e}")
            # Fallback
            text = self.html_to_text(html_content)
            if text:
                paragraphs = text.split('\n\n')
                result = []
                for para in paragraphs:
                    if para.strip():
                        result.append(('paragraph', para.strip()))
                return result
            return []
    
    def create_excel_file(self, problems: List[Dict]):
        """Create Excel file with problem metadata."""
        print("Creating Excel file...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "LeetCode Problems"
        
        # Headers
        headers = ["Problem ID", "Title", "Difficulty", "Category/Tags", "Companies", "Acceptance Rate", "Paid Only"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add problems
        for problem in problems:
            problem_id = problem.get("frontendQuestionId", "N/A")
            title = problem.get("title", "N/A")
            difficulty = problem.get("difficulty", "N/A")
            tags = ", ".join([tag.get("name", "") for tag in problem.get("topicTags", [])])
            companies = problem.get("companies", "")  # Will be set during download_all
            ac_rate = problem.get("acRate", 0)
            paid_only = "Yes" if problem.get("paidOnly", False) else "No"
            
            ws.append([problem_id, title, difficulty, tags, companies, f"{ac_rate:.2f}%", paid_only])
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 12
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save file
        excel_path = os.path.join(self.output_dir, "leetcode_problems.xlsx")
        wb.save(excel_path)
        print(f"Excel file saved: {excel_path}")
    
    def create_pdf_file(self, problem: Dict, problem_detail: Optional[Dict] = None, 
                       solutions: Optional[Dict[str, str]] = None):
        """Create PDF file for a single problem, optionally including solutions."""
        problem_id = problem.get("frontendQuestionId", "Unknown")
        title = problem.get("title", "Unknown")
        title_slug = problem.get("titleSlug", "")
        
        # Sanitize filename
        safe_title = re.sub(r'[^\w\s-]', '', title).strip()
        safe_title = re.sub(r'[-\s]+', '_', safe_title)
        pdf_filename = f"{problem_id}_{safe_title}.pdf"
        pdf_path = os.path.join(self.pdf_dir, pdf_filename)
        
        # Create PDF
        doc = SimpleDocTemplate(pdf_path, pagesize=letter,
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor='#1a1a1a',
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Heading style
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor='#2c3e50',
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Normal style
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            textColor='#333333',
            spaceAfter=12,
            leading=16
        )
        
        # Title
        story.append(Paragraph(f"Problem {problem_id}: {title}", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Basic Information
        story.append(Paragraph("Problem Information", heading_style))
        
        difficulty = problem.get("difficulty", "N/A")
        difficulty_color = {
            "Easy": "#00b8a3",
            "Medium": "#ffc01e",
            "Hard": "#ff375f"
        }.get(difficulty, "#333333")
        
        difficulty_style = ParagraphStyle(
            'Difficulty',
            parent=normal_style,
            textColor=difficulty_color,
            fontSize=12
        )
        
        info_text = f"""
        <b>Difficulty:</b> <font color="{difficulty_color}">{difficulty}</font><br/>
        <b>Acceptance Rate:</b> {problem.get('acRate', 0):.2f}%<br/>
        <b>Paid Only:</b> {'Yes' if problem.get('paidOnly', False) else 'No'}<br/>
        """
        
        tags = problem.get("topicTags", [])
        if tags:
            tag_names = ", ".join([tag.get("name", "") for tag in tags])
            info_text += f"<b>Tags:</b> {tag_names}<br/>"
        
        story.append(Paragraph(info_text, normal_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Problem Description
        if problem_detail:
            content = problem_detail.get("content", "")
            if content:
                story.append(Paragraph("Problem Description", heading_style))
                
                # Parse HTML to structured content
                structured_content = self.parse_html_to_structured_content(content, str(problem_id))
                
                # Code block style
                code_block_style = ParagraphStyle(
                    'CodeBlock',
                    parent=normal_style,
                    fontName='Courier',
                    fontSize=9,
                    leftIndent=20,
                    rightIndent=20,
                    backColor='#f5f5f5',
                    borderPadding=10,
                    spaceAfter=12
                )
                
                # List item style
                list_item_style = ParagraphStyle(
                    'ListItem',
                    parent=normal_style,
                    leftIndent=30,
                    bulletIndent=20,
                    spaceAfter=6
                )
                
                if structured_content:
                    for content_type, content_data in structured_content:
                        if content_type == 'paragraph':
                            if content_data and content_data.strip():
                                if content_data.strip() == '<br/>':
                                    story.append(Spacer(1, 0.1*inch))
                                else:
                                    # Check if this looks like a section heading (Example, Constraints, etc.)
                                    text_lower = content_data.lower()
                                    if (text_lower.startswith('example') or 
                                        text_lower.startswith('constraints') or
                                        text_lower.startswith('follow-up') or
                                        text_lower.startswith('note')):
                                        # Format as section heading
                                        story.append(Paragraph(f"<b>{content_data}</b>", normal_style))
                                        story.append(Spacer(1, 0.15*inch))
                                    else:
                                        # Regular paragraph
                                        safe_text = self._escape_for_reportlab(content_data)
                                        story.append(Paragraph(safe_text, normal_style))
                                        story.append(Spacer(1, 0.1*inch))
                        
                        elif content_type == 'heading':
                            if content_data and content_data.strip():
                                # Format as bold heading
                                story.append(Paragraph(f"<b>{content_data}</b>", normal_style))
                                story.append(Spacer(1, 0.15*inch))
                        
                        elif content_type == 'code':
                            if content_data and content_data.strip():
                                # Code block
                                code_escaped = content_data.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                code_lines = code_escaped.split('\n')
                                code_text = '<br/>'.join(code_lines)
                                story.append(Paragraph(code_text, code_block_style))
                                story.append(Spacer(1, 0.15*inch))
                        
                        elif content_type == 'list_item':
                            if content_data and content_data.strip():
                                story.append(Paragraph(content_data, list_item_style))
                                story.append(Spacer(1, 0.05*inch))
                        
                        elif content_type == 'image':
                            if os.path.exists(content_data) and self.is_supported_image_format(content_data):
                                try:
                                    with PILImage.open(content_data) as img:
                                        img_width, img_height = img.size
                                        max_width = 5 * inch
                                        if img_width > max_width:
                                            scale = max_width / img_width
                                            img_width = max_width
                                            img_height = img_height * scale
                                        
                                        rl_img = RLImage(content_data, width=img_width, height=img_height)
                                        story.append(rl_img)
                                        story.append(Spacer(1, 0.2*inch))
                                except (OSError, IOError):
                                    pass
                                except Exception:
                                    pass
                else:
                    # Fallback
                    description_text = self.html_to_text(content)
                    paragraphs = description_text.split('\n\n')
                    for para in paragraphs:
                        if para.strip():
                            para = self._escape_for_reportlab(para)
                            story.append(Paragraph(para, normal_style))
                            story.append(Spacer(1, 0.1*inch))
    
    def _escape_for_reportlab(self, text: str) -> str:
        """Escape text for ReportLab while preserving ReportLab HTML tags."""
        if not text:
            return ""
        
        # ReportLab supports these tags: b, i, u, font, br, super, sub, para
        # We need to protect these tags while escaping everything else
        
        # First, protect existing HTML entities to avoid double-escaping
        protected_entities = {
            '&amp;': '__ENT_AMP__',
            '&lt;': '__ENT_LT__',
            '&gt;': '__ENT_GT__',
            '&quot;': '__ENT_QUOT__',
            '&#39;': '__ENT_APOS__',
            '&nbsp;': '__ENT_NBSP__',
        }
        
        for entity, placeholder in protected_entities.items():
            text = text.replace(entity, placeholder)
        
        # Use a more robust regex to match ReportLab tags
        # This pattern matches opening/closing tags with optional attributes
        # Handles: <tag>, </tag>, <tag attr="val">, <tag attr='val'>, <br/>, etc.
        allowed_tags = ['b', 'i', 'u', 'font', 'br', 'super', 'sub', 'para']
        tag_pattern = r'(</?(?:' + '|'.join(allowed_tags) + r')(?:\s+[^>]*)?/?>)'
        
        # Find and protect all ReportLab tags
        protected_tags = {}
        tag_counter = 0
        
        def protect_tag(match):
            nonlocal tag_counter
            full_tag = match.group(0)
            placeholder = f"__TAG_{tag_counter}__"
            protected_tags[placeholder] = full_tag
            tag_counter += 1
            return placeholder
        
        # Protect all valid ReportLab tags
        text = re.sub(tag_pattern, protect_tag, text, flags=re.IGNORECASE)
        
        # Now escape all remaining HTML characters
        text = text.replace('&', '&amp;')
        text = text.replace('<', '&lt;')
        text = text.replace('>', '&gt;')
        
        # Restore protected entities
        for entity, placeholder in protected_entities.items():
            text = text.replace(placeholder, entity)
        
        # Restore protected tags (process in order they appeared)
        for placeholder, tag in protected_tags.items():
            text = text.replace(placeholder, tag)
        
        return text
        
        # Code Snippets
        if problem_detail:
            code_snippets = problem_detail.get("codeSnippets", [])
            if code_snippets:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph("Code Snippets", heading_style))
                
                code_style = ParagraphStyle(
                    'Code',
                    parent=normal_style,
                    fontName='Courier',
                    fontSize=9,
                    leftIndent=20,
                    rightIndent=20,
                    backColor='#f5f5f5',
                    borderPadding=10
                )
                
                for snippet in code_snippets:
                    lang = snippet.get("lang", "Unknown")
                    code = snippet.get("code", "")
                    if code:
                        story.append(Paragraph(f"<b>{lang}:</b>", normal_style))
                        # Escape code for ReportLab
                        code_escaped = code.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        # Preserve line breaks
                        code_lines = code_escaped.split('\n')
                        code_text = '<br/>'.join(code_lines)
                        story.append(Paragraph(code_text, code_style))
                        story.append(Spacer(1, 0.1*inch))
        
        # Solutions Section
        if solutions:
            story.append(Spacer(1, 0.2*inch))
            story.append(Paragraph("Solutions", heading_style))
            
            solution_code_style = ParagraphStyle(
                'Code',
                parent=normal_style,
                fontName='Courier',
                fontSize=9,
                leftIndent=20,
                rightIndent=20,
                backColor='#e8f4f8',
                borderPadding=10
            )
            
            for lang, solution_code in solutions.items():
                story.append(Paragraph(f"<b>{lang} Solution:</b>", normal_style))
                # Escape code for ReportLab
                code_escaped = solution_code.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                # Preserve line breaks
                code_lines = code_escaped.split('\n')
                code_text = '<br/>'.join(code_lines)
                story.append(Paragraph(code_text, solution_code_style))
                story.append(Spacer(1, 0.15*inch))
        
        # Build PDF
        try:
            doc.build(story)
            return pdf_path
        except Exception as e:
            print(f"Error creating PDF for {title}: {e}")
            return None
    
    def download_all(self, limit: Optional[int] = None):
        """Main method to download all problems and generate files.
        
        Args:
            limit: Optional limit on number of problems to process (for testing)
        """
        print("Starting LeetCode problem downloader...")
        
        # Fetch all problems
        problems = self.fetch_all_problems()
        
        if not problems:
            print("No problems fetched. Exiting.")
            return

        # Apply limit if specified
        if limit is not None and limit > 0:
            problems = problems[:limit]
            print(f"\n⚠️  TEST MODE: Processing only first {limit} problems")

        # Check how many paid problems were fetched
        paid_count = sum(1 for p in problems if p.get("paidOnly", False))
        print(f"Total problems to process: {len(problems)}")
        print(f"Paid-only problems: {paid_count}")
        print(f"Free problems: {len(problems) - paid_count}")
        
        # Fetch company information for each problem
        print(f"\nFetching company information for {len(problems)} problems...")
        for i, problem in enumerate(problems, 1):
            title_slug = problem.get("titleSlug", "")
            if title_slug and i % 10 == 0:
                print(f"Fetching company info: {i}/{len(problems)}...")
            
            if title_slug:
                problem_detail = self.fetch_problem_detail(title_slug)
                if problem_detail:
                    company_tag_stats = problem_detail.get("companyTagStats")
                    companies = self.extract_companies(company_tag_stats)
                    problem["companies"] = companies
                else:
                    problem["companies"] = ""
                time.sleep(0.2)  # Be respectful to the API
            else:
                problem["companies"] = ""
        
        # Create Excel file
        self.create_excel_file(problems)
        
        # Create PDF files for each problem
        print(f"\nCreating PDF files for {len(problems)} problems...")
        successful = 0
        failed = 0
        
        for i, problem in enumerate(problems, 1):
            title = problem.get("title", "Unknown")
            title_slug = problem.get("titleSlug", "")
            
            print(f"[{i}/{len(problems)}] Processing: {title}")
            
            # Fetch problem details
            problem_detail = None
            if title_slug:
                problem_detail = self.fetch_problem_detail(title_slug)
                time.sleep(0.3)  # Be respectful to the API
            
            # Create PDF
            pdf_path = self.create_pdf_file(problem, problem_detail)
            
            if pdf_path:
                successful += 1
            else:
                failed += 1
            
            # Progress update every 10 problems
            if i % 10 == 0:
                print(f"Progress: {i}/{len(problems)} (Success: {successful}, Failed: {failed})")
        
        print(f"\nCompleted!")
        print(f"Total problems: {len(problems)}")
        print(f"Successful PDFs: {successful}")
        print(f"Failed PDFs: {failed}")
        print(f"\nFiles saved in: {self.output_dir}")
    
    def read_problems_from_excel(self) -> List[Dict]:
        """Read problems from the Excel file."""
        excel_path = os.path.join(self.output_dir, "leetcode_problems.xlsx")
        if not os.path.exists(excel_path):
            print(f"Excel file not found: {excel_path}")
            return []
        
        problems = []
        try:
            wb = load_workbook(excel_path, read_only=True)
            ws = wb.active
            
            # Skip header row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Problem ID and Title exist
                    problem = {
                        'frontendQuestionId': str(row[0]),
                        'title': str(row[1]),
                        'difficulty': str(row[2]) if len(row) > 2 else 'Unknown',
                        'tags': str(row[3]) if len(row) > 3 else '',
                        'titleSlug': self._title_to_slug(str(row[1])),
                    }
                    problems.append(problem)
            
            wb.close()
            print(f"Read {len(problems)} problems from Excel file")
        except Exception as e:
            print(f"Error reading Excel file: {e}")
        
        return problems
    
    def _title_to_slug(self, title: str) -> str:
        """Convert problem title to slug format."""
        # Remove special characters and convert to lowercase
        slug = re.sub(r'[^\w\s-]', '', title.lower())
        slug = re.sub(r'[-\s]+', '-', slug)
        return slug.strip('-')
    
    def get_language_extension(self, language: str) -> str:
        """Get file extension for a programming language."""
        lang_lower = language.lower()
        return self.solution_generator.lang_extensions.get(lang_lower, '.txt')
    
    def save_code_file(self, problem_id: str, title_slug: str, language: str, code: str) -> Optional[str]:
        """Save code solution to a file."""
        try:
            # Create problem directory
            problem_dir = os.path.join(self.codes_dir, f"{problem_id}_{title_slug}")
            os.makedirs(problem_dir, exist_ok=True)
            
            # Get file extension
            ext = self.get_language_extension(language)
            
            # Create filename
            lang_slug = language.lower().replace(' ', '_').replace('#', 'sharp')
            filename = f"{lang_slug}{ext}"
            filepath = os.path.join(problem_dir, filename)
            
            # Save code
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(code)
            
            return filepath
        except Exception as e:
            print(f"Error saving code file: {e}")
            return None
    
    def generate_solution(self, problem_detail: Dict, language: str, starter_code: str) -> str:
        """Generate optimized solution using SolutionGenerator."""
        return self.solution_generator.generate_solution(problem_detail, language, starter_code)
    
    def update_pdf_with_solutions(self, problem: Dict, problem_detail: Optional[Dict], 
                                  solutions: Dict[str, str]) -> bool:
        """Update PDF file to include generated solutions."""
        # Rebuild PDF with solutions
        return self.create_pdf_file(problem, problem_detail, solutions) is not None
    
    def generate_all_solutions(self):
        """Generate solutions for all problems from Excel file."""
        print("Starting solution generation for all problems...")
        
        # Read problems from Excel
        problems = self.read_problems_from_excel()
        
        if not problems:
            print("No problems found in Excel file. Exiting.")
            return
        
        print(f"Found {len(problems)} problems to process")
        
        successful = 0
        failed = 0
        
        for i, problem in enumerate(problems, 1):
            problem_id = problem.get('frontendQuestionId', 'Unknown')
            title = problem.get('title', 'Unknown')
            title_slug = problem.get('titleSlug', '')
            
            print(f"\n[{i}/{len(problems)}] Processing: {title} (ID: {problem_id})")
            
            # Fetch problem details
            problem_detail = None
            if title_slug:
                problem_detail = self.fetch_problem_detail(title_slug)
                time.sleep(0.3)  # Be respectful to the API
            
            if not problem_detail:
                print(f"  Warning: Could not fetch details for {title}")
                failed += 1
                continue
            
            # Get code snippets
            code_snippets = problem_detail.get("codeSnippets", [])
            if not code_snippets:
                print(f"  Warning: No code snippets found for {title}")
                failed += 1
                continue
            
            # Generate solutions for each language
            solutions = {}
            code_files_saved = []
            
            for snippet in code_snippets:
                language = snippet.get("lang", "Unknown")
                starter_code = snippet.get("code", "")
                
                if not starter_code:
                    continue
                
                print(f"  Generating {language} solution...")
                
                # Generate solution
                try:
                    solution_code = self.generate_solution(problem_detail, language, starter_code)
                    
                    # Save code file
                    filepath = self.save_code_file(problem_id, title_slug, language, solution_code)
                    if filepath:
                        code_files_saved.append(filepath)
                        solutions[language] = solution_code
                        print(f"    Saved: {filepath}")
                except Exception as e:
                    print(f"    Error generating {language} solution: {e}")
            
            # Update PDF with solutions
            if solutions:
                print(f"  Updating PDF with solutions...")
                if self.update_pdf_with_solutions(problem, problem_detail, solutions):
                    print(f"    PDF updated successfully")
                    successful += 1
                else:
                    print(f"    Failed to update PDF")
                    failed += 1
            else:
                print(f"  No solutions generated")
                failed += 1
            
            # Progress update
            if i % 10 == 0:
                print(f"\nProgress: {i}/{len(problems)} (Success: {successful}, Failed: {failed})")
        
        print(f"\n{'='*60}")
        print(f"Solution generation completed!")
        print(f"Total problems: {len(problems)}")
        print(f"Successful: {successful}")
        print(f"Failed: {failed}")
        print(f"Code files saved in: {self.codes_dir}")
        print(f"PDFs updated in: {self.pdf_dir}")


def main():
    """Main entry point.
    
    Usage:
        python leetcode_downloader.py                    # Download all problems
        python leetcode_downloader.py --limit 5          # Download first 5 problems (for testing)
        python leetcode_downloader.py -n 5               # Same as above (short form)
        python leetcode_downloader.py generate            # Generate solutions for all problems
    """
    import sys
    import argparse
    
    session_cookie = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJfYXV0aF91c2VyX2lkIjoiMzMzOTE0OCIsIl9hdXRoX3VzZXJfYmFja2VuZCI6ImRqYW5nby5jb250cmliLmF1dGguYmFja2VuZHMuTW9kZWxCYWNrZW5kIiwiX2F1dGhfdXNlcl9oYXNoIjoiNjQ0NDliZmM1MjgzNzQxMjhiZDQzMDU2MGQyMzhkODFiZjY3NDVkZDA1ZmM1NTllZWZjNmRjNWI2MDE5NmM4MyIsInNlc3Npb25fdXVpZCI6IjM3ZWRlYTEyIiwiaWQiOjMzMzkxNDgsImVtYWlsIjoibGFpLm5lbHNvbkBnbWFpbC5jb20iLCJ1c2VybmFtZSI6Ik5lbHNvbllITGFpIiwidXNlcl9zbHVnIjoiTmVsc29uWUhMYWkiLCJhdmF0YXIiOiJodHRwczovL2Fzc2V0cy5sZWV0Y29kZS5jb20vdXNlcnMvZGVmYXVsdF9hdmF0YXIuanBnIiwicmVmcmVzaGVkX2F0IjoxNzYzNTc4MjkyLCJpcCI6IjI2MDc6ZmVhODozYjlmOjlmZTA6NTk1ZjphNjVkOmNhZDA6ODk0NiIsImlkZW50aXR5IjoiZjE4YjUyMTNiNmRlMjQ5MGVjOWJlMjE4YjBmMDI1YjAiLCJkZXZpY2Vfd2l0aF9pcCI6WyJhYTVkMjg4YTRhOTc5MDk0ZDA4ODI4OTgwMDAyYzI1NiIsIjI2MDc6ZmVhODozYjlmOjlmZTA6NTk1ZjphNjVkOmNhZDA6ODk0NiJdLCJfc2Vzc2lvbl9leHBpcnkiOjEyMDk2MDB9.ULoZrT6j02RIrpLRfmdVluUqMfWd_F8ZHFPTaaMiWu4"  # Remove before committing!
    downloader = LeetCodeDownloader(session_cookie=session_cookie)
    
    # Check for 'generate' command (old syntax)
    if len(sys.argv) > 1 and sys.argv[1] == 'generate':
        # Old syntax: python leetcode_downloader.py generate
        downloader.generate_all_solutions()
    else:
        # Parse arguments for download command
        parser = argparse.ArgumentParser(description='LeetCode Problem Downloader')
        parser.add_argument('--limit', '-n', type=int, default=None,
                           help='Limit the number of problems to process (for testing)')
        args = parser.parse_args()
        
        # Download problems (with optional limit)
        downloader.download_all(limit=args.limit)

if __name__ == "__main__":
    main()

